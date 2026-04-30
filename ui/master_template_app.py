"""
Master Template Generator
Hanger + Stack Bulk Upload → Final Master Template (StockTake Template)
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Master Template Generator",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@500;700&display=swap');

    /* ══════════════════════════════════════════════════════════════════════
       KEYFRAME ANIMATIONS
       ══════════════════════════════════════════════════════════════════════ */
    @keyframes gradientShift {
        0%   { background-position: 0% 50%; }
        50%  { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(14px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes fadeInScale {
        from { opacity: 0; transform: scale(0.96); }
        to   { opacity: 1; transform: scale(1); }
    }
    @keyframes float {
        0%, 100% { transform: translateY(0); }
        50%      { transform: translateY(-3px); }
    }
    @keyframes pulse {
        0%, 100% { box-shadow: 0 0 0 0 rgba(15, 98, 254, 0.45); }
        50%      { box-shadow: 0 0 0 12px rgba(15, 98, 254, 0); }
    }
    @keyframes shimmer {
        0%   { background-position: -200% 0; }
        100% { background-position: 200% 0; }
    }
    @keyframes spinSlow {
        from { transform: rotate(0deg); }
        to   { transform: rotate(360deg); }
    }
    @keyframes glowBorder {
        0%, 100% { border-color: rgba(15, 98, 254, 0.25); box-shadow: 0 0 14px rgba(15, 98, 254, 0.08); }
        50%      { border-color: rgba(105, 41, 196, 0.55);  box-shadow: 0 0 22px rgba(105, 41, 196, 0.18); }
    }
    @keyframes blob {
        0%, 100% { transform: translate(0, 0) scale(1); }
        33%      { transform: translate(40px, -30px) scale(1.1); }
        66%      { transform: translate(-30px, 30px) scale(0.95); }
    }

    /* ══════════════════════════════════════════════════════════════════════
       BASE TYPOGRAPHY + APP BACKGROUND
       ══════════════════════════════════════════════════════════════════════ */
    html, body, [class*="css"], .stApp, .stMarkdown, .stText {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }
    code, pre, .stCode { font-family: 'JetBrains Mono', monospace !important; }

    .stApp {
        background:
            radial-gradient(1200px 600px at 0% 0%, rgba(15, 98, 254, 0.08), transparent 60%),
            radial-gradient(1000px 500px at 100% 0%, rgba(105, 41, 196, 0.07), transparent 60%),
            radial-gradient(900px 500px at 50% 100%, rgba(8, 189, 186, 0.06), transparent 60%),
            linear-gradient(180deg, #f4f6fb 0%, #eef1f8 100%);
        background-attachment: fixed;
    }
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 4rem;
        max-width: 1400px;
        animation: fadeInUp 0.6s ease-out;
    }

    /* ══════════════════════════════════════════════════════════════════════
       MAIN HEADER — animated gradient text + floating accent
       ══════════════════════════════════════════════════════════════════════ */
    .main-header {
        font-size: 2.8rem;
        font-weight: 900;
        letter-spacing: -1px;
        background: linear-gradient(110deg, #0F62FE 0%, #6929C4 30%, #FF4D8D 55%, #08BDBA 80%, #0F62FE 100%);
        background-size: 300% 300%;
        animation: gradientShift 8s ease-in-out infinite, fadeInUp 0.7s ease-out;
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        text-align: center;
        padding: 0.5rem 0 0;
        margin: 0;
        line-height: 1.15;
    }
    .main-subtitle {
        text-align: center;
        color: #5b6478;
        font-size: 1rem;
        font-weight: 500;
        margin: 0.4rem 0 1.75rem;
        letter-spacing: 0.2px;
        animation: fadeInUp 0.9s ease-out;
    }
    .main-subtitle::before, .main-subtitle::after {
        content: '';
        display: inline-block;
        width: 36px;
        height: 2px;
        background: linear-gradient(90deg, transparent, #0F62FE);
        vertical-align: middle;
        margin: 0 12px;
        border-radius: 2px;
    }
    .main-subtitle::after {
        background: linear-gradient(90deg, #6929C4, transparent);
    }

    /* ══════════════════════════════════════════════════════════════════════
       SECTION BANNER — animated gradient bg + shimmer overlay
       ══════════════════════════════════════════════════════════════════════ */
    .section-banner {
        position: relative;
        background: linear-gradient(135deg, #0d0d1f 0%, #1a1a2e 40%, #16213e 100%);
        background-size: 200% 200%;
        animation: gradientShift 12s ease infinite, fadeInScale 0.5s ease-out;
        border-radius: 16px;
        padding: 1.4rem 1.9rem;
        margin: 1.25rem 0 1.5rem;
        box-shadow: 0 10px 30px rgba(15, 23, 42, 0.25), 0 1px 0 rgba(255,255,255,0.04) inset;
        border: 1px solid rgba(255,255,255,0.06);
        overflow: hidden;
    }
    .section-banner::before {
        content: '';
        position: absolute;
        inset: 0;
        background: linear-gradient(110deg, transparent 30%, rgba(255,255,255,0.07) 50%, transparent 70%);
        background-size: 200% 100%;
        animation: shimmer 6s linear infinite;
        pointer-events: none;
    }
    .section-banner::after {
        content: '';
        position: absolute;
        top: -40px; right: -40px;
        width: 160px; height: 160px;
        background: radial-gradient(circle, rgba(15, 98, 254, 0.35), transparent 70%);
        border-radius: 50%;
        animation: blob 10s ease-in-out infinite;
        pointer-events: none;
    }
    .section-banner .banner-title {
        position: relative;
        font-size: 1.55rem;
        font-weight: 800;
        color: #ffffff;
        line-height: 1.2;
        margin: 0;
        letter-spacing: -0.3px;
        text-shadow: 0 2px 12px rgba(0,0,0,0.3);
    }
    .section-banner .banner-sub {
        position: relative;
        font-size: 0.9rem;
        color: #c8ccdf;
        margin: 0.4rem 0 0;
        line-height: 1.5;
    }
    .section-banner.accent-blue   { background: linear-gradient(135deg, #003a8c 0%, #0F62FE 50%, #003a8c 100%); background-size: 200% 200%; }
    .section-banner.accent-green  { background: linear-gradient(135deg, #044317 0%, #198038 50%, #044317 100%); background-size: 200% 200%; }
    .section-banner.accent-purple { background: linear-gradient(135deg, #31135E 0%, #6929C4 50%, #31135E 100%); background-size: 200% 200%; }

    /* ══════════════════════════════════════════════════════════════════════
       UPLOAD CARD — gradient + lift on hover + animated border ribbon
       ══════════════════════════════════════════════════════════════════════ */
    .upload-card {
        position: relative;
        background: linear-gradient(135deg, #0F62FE 0%, #6929C4 50%, #FF4D8D 100%);
        background-size: 200% 200%;
        animation: gradientShift 10s ease infinite, fadeInUp 0.6s ease-out;
        padding: 1.35rem 1.6rem;
        border-radius: 16px;
        color: white;
        margin-bottom: 0.85rem;
        box-shadow: 0 10px 28px rgba(15, 98, 254, 0.28);
        overflow: hidden;
        transition: transform 0.25s cubic-bezier(0.4, 0, 0.2, 1), box-shadow 0.25s ease;
        cursor: default;
    }
    .upload-card::before {
        content: '';
        position: absolute;
        top: 0; left: -100%;
        width: 100%; height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.18), transparent);
        transition: left 0.6s ease;
    }
    .upload-card:hover {
        transform: translateY(-4px) scale(1.01);
        box-shadow: 0 16px 40px rgba(15, 98, 254, 0.4);
    }
    .upload-card:hover::before { left: 100%; }
    .upload-card .upload-title {
        position: relative;
        font-size: 1.15rem;
        font-weight: 800;
        color: #ffffff;
        margin: 0 0 0.3rem;
        line-height: 1.25;
        letter-spacing: 0.1px;
        text-shadow: 0 1px 6px rgba(0,0,0,0.18);
    }
    .upload-card .upload-sub {
        position: relative;
        font-size: 0.85rem;
        color: rgba(255,255,255,0.93);
        margin: 0;
        line-height: 1.45;
    }

    /* ══════════════════════════════════════════════════════════════════════
       STATUS CARDS — success / warn / info with glow + lift
       ══════════════════════════════════════════════════════════════════════ */
    .success-card, .warn-card, .info-card {
        position: relative;
        padding: 1.1rem 1.35rem;
        border-radius: 14px;
        color: white;
        margin: 0.6rem 0;
        overflow: hidden;
        animation: fadeInScale 0.45s ease-out;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .success-card:hover, .warn-card:hover, .info-card:hover { transform: translateY(-2px); }
    .success-card {
        background: linear-gradient(135deg, #0e7c75 0%, #11998e 50%, #38ef7d 100%);
        background-size: 200% 200%;
        animation: gradientShift 8s ease infinite, fadeInScale 0.45s ease-out;
        box-shadow: 0 8px 22px rgba(17, 153, 142, 0.35);
    }
    .success-card:hover { box-shadow: 0 12px 32px rgba(17, 153, 142, 0.5); }
    .warn-card {
        background: linear-gradient(135deg, #f5576c 0%, #f093fb 100%);
        background-size: 200% 200%;
        animation: gradientShift 8s ease infinite, fadeInScale 0.45s ease-out;
        box-shadow: 0 8px 22px rgba(245, 87, 108, 0.35);
    }
    .warn-card:hover { box-shadow: 0 12px 32px rgba(245, 87, 108, 0.5); }
    .info-card {
        background: linear-gradient(135deg, #3494E6 0%, #5e60ce 100%);
        background-size: 200% 200%;
        animation: gradientShift 8s ease infinite, fadeInScale 0.45s ease-out;
        box-shadow: 0 8px 22px rgba(52, 148, 230, 0.35);
    }
    .info-card:hover { box-shadow: 0 12px 32px rgba(52, 148, 230, 0.5); }
    .success-card .card-title, .warn-card .card-title, .info-card .card-title {
        font-size: 1rem; font-weight: 800; margin: 0 0 0.25rem; color: #fff;
    }
    .success-card .card-body, .warn-card .card-body, .info-card .card-body {
        font-size: 0.88rem; color: rgba(255,255,255,0.94); margin: 0; line-height: 1.45;
    }

    /* ══════════════════════════════════════════════════════════════════════
       METRIC CARDS — glassmorphism + gradient number + hover scale
       ══════════════════════════════════════════════════════════════════════ */
    .metric-box {
        position: relative;
        background: rgba(255, 255, 255, 0.85);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        padding: 1.1rem 1.3rem;
        border-radius: 14px;
        box-shadow: 0 4px 16px rgba(15, 23, 42, 0.06);
        border: 1px solid rgba(230, 233, 240, 0.8);
        text-align: center;
        transition: transform 0.25s cubic-bezier(0.4, 0, 0.2, 1), box-shadow 0.25s ease, border-color 0.25s ease;
        animation: fadeInUp 0.5s ease-out;
        overflow: hidden;
    }
    .metric-box::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, #0F62FE, #6929C4, #08BDBA);
        background-size: 200% 100%;
        animation: gradientShift 4s linear infinite;
        opacity: 0;
        transition: opacity 0.25s ease;
    }
    .metric-box:hover {
        transform: translateY(-4px) scale(1.02);
        box-shadow: 0 12px 28px rgba(15, 98, 254, 0.15);
        border-color: rgba(15, 98, 254, 0.3);
    }
    .metric-box:hover::before { opacity: 1; }
    .metric-num {
        font-size: 2rem;
        font-weight: 900;
        background: linear-gradient(135deg, #0F62FE, #6929C4);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        line-height: 1.1;
        letter-spacing: -0.5px;
    }
    .metric-lbl {
        font-size: 0.78rem;
        color: #5b6478;
        margin-top: 0.3rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.6px;
    }

    /* ══════════════════════════════════════════════════════════════════════
       STREAMLIT NATIVE — readability + interaction polish
       ══════════════════════════════════════════════════════════════════════ */
    h1, h2, h3, h4, h5, h6 { color: #1a1a2e !important; font-weight: 700 !important; letter-spacing: -0.2px; }
    .stMarkdown p { color: #2d3142; }

    /* File uploader */
    [data-testid="stFileUploader"] section {
        background: #ffffff;
        border: 2px dashed #c7cdd9;
        border-radius: 14px;
        padding: 1.1rem;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: #0F62FE;
        background: linear-gradient(135deg, #ffffff, #f0f5ff);
        transform: translateY(-1px);
        box-shadow: 0 8px 22px rgba(15, 98, 254, 0.12);
    }
    [data-testid="stFileUploader"] section::before {
        content: '';
        position: absolute;
        inset: 0;
        background: radial-gradient(circle at var(--mx, 50%) var(--my, 50%), rgba(15, 98, 254, 0.08), transparent 40%);
        opacity: 0;
        transition: opacity 0.3s ease;
        pointer-events: none;
    }
    [data-testid="stFileUploader"] section:hover::before { opacity: 1; }

    /* Buttons — animated gradient + lift + shimmer */
    .stButton > button {
        position: relative;
        background: linear-gradient(135deg, #0F62FE 0%, #6929C4 50%, #FF4D8D 100%);
        background-size: 200% 200%;
        color: white !important;
        border: none;
        border-radius: 12px;
        padding: 0.6rem 1.6rem;
        font-weight: 700;
        font-size: 0.95rem;
        letter-spacing: 0.2px;
        box-shadow: 0 6px 18px rgba(15, 98, 254, 0.28);
        transition: transform 0.2s cubic-bezier(0.4, 0, 0.2, 1), box-shadow 0.2s ease, background-position 0.4s ease;
        overflow: hidden;
    }
    .stButton > button:hover {
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 12px 28px rgba(105, 41, 196, 0.4);
        background-position: 100% 100%;
    }
    .stButton > button:active { transform: translateY(0) scale(0.99); }

    /* Download button — green ramp */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #0e6027 0%, #198038 50%, #38ef7d 100%);
        background-size: 200% 200%;
        color: white !important;
        border: none;
        border-radius: 12px;
        font-weight: 700;
        box-shadow: 0 6px 18px rgba(25, 128, 56, 0.3);
        transition: transform 0.2s ease, box-shadow 0.2s ease, background-position 0.4s ease;
        animation: pulse 2.4s ease-in-out infinite;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 12px 28px rgba(56, 239, 125, 0.45);
        background-position: 100% 100%;
        animation: none;
    }

    /* Streamlit metrics */
    [data-testid="stMetric"] {
        background: rgba(255, 255, 255, 0.7);
        backdrop-filter: blur(10px);
        border-radius: 12px;
        padding: 0.9rem 1rem;
        border: 1px solid rgba(230, 233, 240, 0.8);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    [data-testid="stMetric"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(15, 98, 254, 0.1);
    }
    [data-testid="stMetricValue"] {
        font-weight: 900;
        background: linear-gradient(135deg, #0F62FE, #6929C4);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    [data-testid="stMetricLabel"] {
        font-weight: 600;
        color: #5b6478;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-size: 0.78rem !important;
    }

    /* Tabs — animated underline + gradient active state */
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
        border-bottom: 2px solid #e6e9f0;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 10px 10px 0 0;
        padding: 0.6rem 1.2rem;
        font-weight: 600;
        color: #5b6478;
        transition: color 0.2s ease, background 0.2s ease;
        position: relative;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(15, 98, 254, 0.06);
        color: #0F62FE;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #0F62FE 0%, #6929C4 100%) !important;
        color: white !important;
        box-shadow: 0 4px 14px rgba(15, 98, 254, 0.3);
    }

    /* Expanders */
    div[data-testid="stExpander"] {
        border: 1px solid #e6e9f0;
        border-radius: 14px;
        background: #ffffff;
        box-shadow: 0 2px 10px rgba(15, 23, 42, 0.04);
        transition: box-shadow 0.2s ease, border-color 0.2s ease;
        animation: fadeInUp 0.4s ease-out;
    }
    div[data-testid="stExpander"]:hover {
        border-color: rgba(15, 98, 254, 0.3);
        box-shadow: 0 6px 20px rgba(15, 98, 254, 0.08);
    }

    /* Dataframes */
    .stDataFrame, [data-testid="stDataFrame"] {
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
        border: 1px solid #e6e9f0;
        animation: fadeInUp 0.5s ease-out;
    }

    /* Alerts (st.info / st.warning / st.success / st.error) */
    [data-testid="stAlert"] {
        border-radius: 12px;
        border: none;
        box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
        animation: fadeInUp 0.4s ease-out;
    }

    /* Spinners — slower, smoother */
    .stSpinner > div { border-top-color: #0F62FE !important; }

    /* Sidebar polish */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #ffffff 0%, #f4f6fb 100%);
        border-right: 1px solid #e6e9f0;
    }

    /* Progress bar — animated gradient fill */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #0F62FE, #6929C4, #08BDBA, #0F62FE);
        background-size: 200% 100%;
        animation: gradientShift 3s linear infinite;
    }

    /* Plotly chart container — subtle entrance */
    [data-testid="stPlotlyChart"] {
        animation: fadeInUp 0.6s ease-out;
        border-radius: 14px;
        overflow: hidden;
    }

    /* Selectbox / text input focus glow */
    [data-baseweb="select"] > div, [data-baseweb="input"] > div {
        border-radius: 10px !important;
        transition: box-shadow 0.2s ease, border-color 0.2s ease;
    }
    [data-baseweb="select"] > div:focus-within, [data-baseweb="input"] > div:focus-within {
        box-shadow: 0 0 0 3px rgba(15, 98, 254, 0.15) !important;
        border-color: #0F62FE !important;
    }

    /* Hide deploy button & footer for cleaner look */
    [data-testid="stToolbar"] { visibility: hidden; }
    footer { visibility: hidden; }
    #MainMenu { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ── Currency conversion table (from master template) ──────────────────────────
CURRENCY_CONVERSION = {
    'United Arab Emirates': 1.0,
    'Qatar': 1.0,
    'Netherlands': 2.04,
    'Oman': 9.54,
    'Mauritius': 0.08,
    'Kuwait': 12.02,
    'Israel': 1.0,
    'Bahrain': 9.74,
    'Kenya': 0.028,
    'Malta': 9.48,
    'Lebanon': 3.67,
    'Singapore': 2.82,
    'South Africa': 0.2,
    'Kingdom of Saudi Arabia': 0.98,
    'Australia': 2.28,
}

# ── Low Value file detector ───────────────────────────────────────────────────
def is_low_value_file(file) -> bool:
    """Returns True if the uploaded file is a Low Value stocktake file.
    Detection rules (both must be true):
      1. Has Sheet2 and/or Sheet3 (LV files carry their own lookup sheets)
      2. Sheet1 does NOT have a 'Room' column (regular totes/hanger files always have 'Room')
    """
    try:
        file.seek(0)
        raw = file.read()
        file.seek(0)
        xl = pd.ExcelFile(BytesIO(raw), engine='openpyxl')
        has_extra_sheets = 'Sheet2' in xl.sheet_names or 'Sheet3' in xl.sheet_names
        if not has_extra_sheets:
            return False
        # Check Sheet1 columns — if 'Room' exists it's a regular totes/hanger file
        s1 = pd.read_excel(BytesIO(raw), sheet_name='Sheet1', engine='openpyxl', nrows=2)
        if 'Room' in s1.columns:
            return False
        return True
    except Exception:
        file.seek(0)
        return False


# ── Stack Bulk reader with auto-repair ────────────────────────────────────────
def read_stack_bulk(file):
    """
    Read Stack Bulk Upload file (BulkSell sheet).
    Tries multiple methods to handle corrupted / validation-locked files.
    Returns (dataframe, warning_message_or_None).
    """
    file.seek(0)
    raw = file.read()

    sheet_candidates = ['BulkSell', 0]

    # Method 1 — standard openpyxl
    for sheet in sheet_candidates:
        try:
            df = pd.read_excel(BytesIO(raw), sheet_name=sheet, engine='openpyxl')
            if len(df) > 0:
                return df, None
        except Exception:
            pass

    # Method 2 — data_only=True (skips formula errors)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, keep_links=False)
        sheet_name = 'BulkSell' if 'BulkSell' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        df = pd.DataFrame(data, columns=headers)
        if len(df) > 0:
            return df, "File repaired: data validation rules removed."
    except Exception:
        pass

    # Method 3 — strip dataValidation XML tags directly from the ZIP then re-read
    try:
        import zipfile, re as _re
        zin  = zipfile.ZipFile(BytesIO(raw))
        buf2 = BytesIO()
        zout = zipfile.ZipFile(buf2, 'w', zipfile.ZIP_DEFLATED)
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/'):
                data = _re.sub(
                    rb'<dataValidations[^>]*>.*?</dataValidations>', b'',
                    data, flags=_re.DOTALL
                )
            zout.writestr(item, data)
        zout.close()
        buf2.seek(0)
        df = pd.read_excel(buf2, sheet_name='BulkSell', engine='openpyxl')
        if len(df) > 0:
            return df, "File repaired: invalid data validation rules stripped."
    except Exception:
        pass

    # Method 4 — xlrd (older .xls formats)
    try:
        df = pd.read_excel(BytesIO(raw), sheet_name=0, engine='xlrd')
        if len(df) > 0:
            return df, "File read using xlrd engine."
    except Exception:
        pass

    return None, "ERROR"


# ── Hanger processing ─────────────────────────────────────────────────────────
def process_hanger(file):
    """
    1. Filter AE: Deal ID → BACKEND → 'No deal ID'
    2. Merge location: Hangar + LOA  →  e.g. '61A'
    3. Return DataFrame with: Room, Bin, Location, IMEI, Deal Id
    """
    file.seek(0)
    df = pd.read_excel(file, sheet_name='Sheet1', engine='openpyxl')

    # Drop completely empty rows — keep only rows where Room is filled
    df = df[df['Room'].notna()].reset_index(drop=True)
    original_rows = len(df)

    # Step 1 — AE in Deal ID
    df['_deal'] = df['Deal ID'].where(
        df['Deal ID'].astype(str).str.contains('AE', na=False, case=False)
    )
    ae_deal_id = int(df['_deal'].notna().sum())

    # Step 2 — fill blanks from BACKEND where BACKEND contains AE
    mask = df['BACKEND'].astype(str).str.contains('AE', na=False, case=False)
    df.loc[df['_deal'].isna() & mask, '_deal'] = \
        df.loc[df['_deal'].isna() & mask, 'BACKEND']
    ae_backend = int(df['_deal'].notna().sum()) - ae_deal_id

    # Step 3 — remaining blanks → 'No deal ID'
    df['_deal'] = df['_deal'].fillna('No deal ID')
    no_deal = int((df['_deal'] == 'No deal ID').sum())

    # Merge location: Hangar number + LOA letter
    df['_location'] = (
        df['Hangar'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
        + df['LOA'].astype(str).str.strip()
    )

    # Convert IMEI to clean string — integers come in as int dtype and would
    # render as scientific notation in Excel without this conversion
    def clean_imei(val):
        if pd.isna(val):
            return ''
        if isinstance(val, float):
            return str(int(val))
        return str(val).strip()

    result = pd.DataFrame({
        'Room':     df['Room'],
        'Bin':      'Hanger',
        'Location': df['_location'],
        'IMEI':     df['IMEI'].apply(clean_imei),
        'Deal Id':  df['_deal'],
    }).reset_index(drop=True)

    stats = {
        'original_rows': original_rows,
        'ae_deal_id':    ae_deal_id,
        'ae_backend':    ae_backend,
        'no_deal_id':    no_deal,
    }
    return result, stats


# ── Totes processing ──────────────────────────────────────────────────────────
def process_totes(file):
    """
    Same AE filter as hanger but NO location merge.
    Location = LOA letter as-is.
    Bin = 'Totes'
    """
    file.seek(0)
    df = pd.read_excel(file, sheet_name='Sheet1', engine='openpyxl')

    # Drop completely empty rows — keep only rows where Room is filled
    df = df[df['Room'].notna()].reset_index(drop=True)
    original_rows = len(df)

    # Step 1 — AE in Deal ID
    df['_deal'] = df['Deal ID'].where(
        df['Deal ID'].astype(str).str.contains('AE', na=False, case=False)
    )
    ae_deal_id = int(df['_deal'].notna().sum())

    # Step 2 — fill blanks from BACKEND where BACKEND contains AE
    mask = df['BACKEND'].astype(str).str.contains('AE', na=False, case=False)
    df.loc[df['_deal'].isna() & mask, '_deal'] = \
        df.loc[df['_deal'].isna() & mask, 'BACKEND']
    ae_backend = int(df['_deal'].notna().sum()) - ae_deal_id

    # Step 3 — remaining blanks → 'No deal ID'
    df['_deal'] = df['_deal'].fillna('No deal ID')
    no_deal = int((df['_deal'] == 'No deal ID').sum())

    # Location = LOA as-is (no merge with Hangar)
    def clean_imei(val):
        if pd.isna(val):
            return ''
        if isinstance(val, float):
            return str(int(val))
        return str(val).strip()

    result = pd.DataFrame({
        'Room':     df['Room'],
        'Bin':      'Totes',
        'Location': df['LOA'].astype(str).str.strip(),
        'IMEI':     df['IMEI'].apply(clean_imei),
        'Deal Id':  df['_deal'],
    }).reset_index(drop=True)

    stats = {
        'original_rows': original_rows,
        'ae_deal_id':    ae_deal_id,
        'ae_backend':    ae_backend,
        'no_deal_id':    no_deal,
    }
    return result, stats


# ── Low Value master template builder ────────────────────────────────────────
def build_low_value_template(lv_file) -> pd.DataFrame:
    """
    Reads a Stack-Bulk-format file and generates the low value master template:
      - Room = 'Inventory'
      - Bin  = 'Totes'
      - IMEI = IMEI Number if present, else Barcode
      - Filters to LVIN1/LVIN2/R151 rows if those codes exist in the file,
        otherwise uses all rows.
    """
    LOW_VALUE_CODES = {'LVIN1', 'LVIN2', 'R151'}

    df, _ = read_stack_bulk(lv_file)
    if df is None:
        raise ValueError("Could not read the Low Value file.")

    # Exclude TV and Home Appliances (stored externally)
    df = df[~df['Category'].astype(str).str.strip().str.lower().isin(['tv', 'home appliances'])].copy()

    # Filter to low value codes if present, else use full file
    qr = df['QR Code Description'].astype(str).str.upper().str.strip()
    lv_mask = qr.isin(LOW_VALUE_CODES)
    if lv_mask.any():
        df = df[lv_mask].copy()

    def _clean(val):
        s = str(val).strip()
        if s.lower() in ('', 'nan', 'none'):
            return ''
        return s.split('.')[0]  # remove decimal from floats

    rows = []
    for _, row in df.iterrows():
        imei_raw  = row.get('IMEI Number', '')
        barcode   = _clean(row.get('Barcode', ''))
        imei_clean = _clean(imei_raw)
        imei_val  = imei_clean if imei_clean else barcode  # fallback to barcode

        country = str(row.get('Storage Member Country', '')).strip()
        conversion = CURRENCY_CONVERSION.get(country, '')

        rows.append({
            'Room':       'Inventory',
            'Bin':        'Totes',
            'Location':   str(row.get('QR Code Description', '')).strip(),
            'IMEI':       imei_val,
            'Deal Id':    str(row.get('Appraisal', '')).strip(),
            'Category':   row.get('Category', ''),
            'Brand':      row.get('Brand', ''),
            'Model':      row.get('Asset Label', ''),
            'Grade':      str(row.get('Sell Grade', '')).strip().rstrip('.').strip(),
            'VAT Type':   row.get('Appraisal VATType', ''),
            'Status':     row.get('Appraisal Status', ''),
            'Stack':      row.get('Existing stack Id & Dealer', ''),
            'Country':    country,
            'PP':         '',
            'Conversion': conversion,
            'PP in AED':  '',
        })
    return pd.DataFrame(rows)


# ── Low Value master template builder ────────────────────────────────────────
def build_low_value_template(lv_file) -> pd.DataFrame:
    """
    Reads the low value stocktake xlsm file:
      Sheet1 — stocktake rows (Deal ID, Barcode, IMEI, Bin No., Bin, Country, Price Purcashed)
      Sheet2 — lookup (Appraisal Code → Label, Brand name, Category Name, Appraisal Status, PP, CON, PP_AED)
      Sheet3 — conversion rates (Country → Conversion rate to AED)

    Output master template:
      Room = 'Inventory', Bin = 'Totes', Location = Bin No.,
      IMEI = IMEI if present else Barcode,
      Deal Id, Category, Brand, Model, Grade, VAT Type, Status, Stack,
      Country, PP, Conversion, PP in AED
    """
    lv_file.seek(0)
    raw = lv_file.read()

    def _clean(val):
        s = str(val).strip()
        return '' if s.lower() in ('', 'nan', 'none') else s.split('.')[0]

    try:
        s1 = pd.read_excel(BytesIO(raw), sheet_name='Sheet1', engine='openpyxl')
        s2 = pd.read_excel(BytesIO(raw), sheet_name='Sheet2', engine='openpyxl')
        s3 = pd.read_excel(BytesIO(raw), sheet_name='Sheet3', engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Could not read Low Value file sheets: {e}")

    # Build conversion rate dict from Sheet3
    conv_map = {}
    for _, r in s3.iterrows():
        c = str(r.iloc[0]).strip()
        try:
            conv_map[c] = float(r.iloc[1])
        except Exception:
            pass

    # Build lookup dict from Sheet2: Appraisal Code → dict
    lookup2 = {}
    for _, r in s2.iterrows():
        key = str(r.get('Appraisal Code', '')).strip()
        if key and key.lower() != 'nan':
            pp_val  = r.get('PP', '')
            con_val = r.get('CON', '')
            try:
                pp_f  = float(pp_val)  if pd.notna(pp_val)  else ''
                con_f = float(con_val) if pd.notna(con_val) else ''
                pp_aed = round(pp_f * con_f, 2) if pp_f != '' and con_f != '' else ''
            except Exception:
                pp_f = pp_aed = ''
            lookup2[key] = {
                'Category': str(r.get('Category Name', '')).strip(),
                'Brand':    str(r.get('Brand name', '')).strip(),
                'Model':    str(r.get('Label', '')).strip(),
                'Status':   str(r.get('Appraisal Status', '')).strip(),
                'Country':  str(r.get('Country', '')).strip(),
                'PP':       pp_f,
                'PP in AED': pp_aed,
            }

    rows = []
    for _, row in s1.iterrows():
        deal_id  = str(row.get('Deal ID', '')).strip()
        imei_raw = _clean(row.get('IMEI', ''))
        barcode  = _clean(row.get('Barcode', ''))
        imei_val = imei_raw if imei_raw else barcode  # fallback to Barcode

        country    = str(row.get('Country', '')).strip()
        conversion = conv_map.get(country, CURRENCY_CONVERSION.get(country, ''))

        s  = lookup2.get(deal_id, {})
        pp = s.get('PP', '') if s else str(row.get('Price Purcashed', '')).strip()
        try:
            pp_num  = float(pp) if pp != '' else ''
            pp_aed  = round(pp_num * conversion, 2) if pp_num != '' and conversion != '' else ''
        except Exception:
            pp_aed = ''

        # Try multiple possible column names for the location/bin field
        _loc = (
            _clean(row.get('Bin No.', '')) or
            _clean(row.get('LOA', ''))     or
            _clean(row.get('Location', '')) or
            _clean(row.get('Bin', ''))
        )

        rows.append({
            'Room':       'Inventory',
            'Bin':        'Totes',
            'Location':   _loc,
            'IMEI':       imei_val,
            'Deal Id':    deal_id,
            'Category':   s.get('Category', str(row.get('Bin', '')).strip()) if s else str(row.get('Bin', '')).strip(),
            'Brand':      s.get('Brand', ''),
            'Model':      s.get('Model', ''),
            'Grade':      '',
            'VAT Type':   '',
            'Status':     s.get('Status', str(row.get('Purchase status', '')).strip()) if s else str(row.get('Purchase status', '')).strip(),
            'Stack':      '',
            'Country':    country,
            'PP':         pp,
            'Conversion': conversion,
            'PP in AED':  s.get('PP in AED', pp_aed) if s else pp_aed,
        })
    return pd.DataFrame(rows)


# ── Stack Bulk lookup builder ──────────────────────────────────────────────────
def build_lookup(stack_df: pd.DataFrame) -> dict:
    """
    Build { appraisal_code → { Category, Brand, Model, Grade,
                                VAT Type, Status, Stack, Country } }
    """
    lookup = {}
    for _, row in stack_df.iterrows():
        key = str(row.get('Appraisal', '')).strip()
        if key and key.lower() != 'nan':
            lookup[key] = {
                'Category': row.get('Category', ''),
                'Brand':    row.get('Brand', ''),
                'Model':    row.get('Asset Label', ''),
                'Grade':    str(row.get('Sell Grade', '')).strip().rstrip('.').strip(),
                'VAT Type': row.get('Appraisal VATType', ''),
                'Status':   row.get('Appraisal Status', ''),
                'Stack':    row.get('Existing stack Id & Dealer', ''),
                'Country':  row.get('Storage Member Country', ''),
            }
    return lookup


# ── Master template builder ───────────────────────────────────────────────────
def build_master_template(hanger_df: pd.DataFrame, lookup: dict) -> pd.DataFrame:
    rows = []
    for _, row in hanger_df.iterrows():
        deal_id = str(row['Deal Id']).strip()
        s = lookup.get(deal_id, {})
        country = str(s.get('Country', '')).strip()
        conversion = CURRENCY_CONVERSION.get(country, '')
        rows.append({
            'Room':       row['Room'],
            'Bin':        row['Bin'],
            'Location':   row['Location'],
            'IMEI':       row['IMEI'],
            'Deal Id':    row['Deal Id'],
            'Category':   s.get('Category', ''),
            'Brand':      s.get('Brand', ''),
            'Model':      s.get('Model', ''),
            'Grade':      s.get('Grade', ''),
            'VAT Type':   s.get('VAT Type', ''),
            'Status':     s.get('Status', ''),
            'Stack':      s.get('Stack', ''),
            'Country':    country,
            'PP':         '',
            'Conversion': conversion,
            'PP in AED':  '',
        })
    return pd.DataFrame(rows)


# ── Excel export with formatting ──────────────────────────────────────────────
def export_excel(df: pd.DataFrame) -> BytesIO:
    import xlsxwriter as xw

    col_widths = {
        'Room': 8, 'Bin': 6, 'Location': 12, 'IMEI': 22,
        'Deal Id': 26, 'Category': 15, 'Brand': 14, 'Model': 32,
        'Grade': 10, 'VAT Type': 18, 'Status': 14, 'Stack': 35,
        'Country': 22, 'PP': 10, 'Conversion': 12, 'PP in AED': 13,
    }

    # Clean IMEI before writing
    df = df.copy()
    if 'IMEI' in df.columns:
        df['IMEI'] = df['IMEI'].apply(
            lambda v: str(v).strip().split('.')[0] if pd.notna(v) else ''
        )

    output = BytesIO()
    wb = xw.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('StockTake Template')

    hdr_fmt = wb.add_format({
        'bold': True, 'font_color': 'white', 'bg_color': '4472C4',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1
    })
    str_fmt  = wb.add_format({'num_format': '@', 'border': 1, 'valign': 'vcenter'})
    cell_fmt = wb.add_format({'border': 1, 'valign': 'vcenter'})

    cols = list(df.columns)
    imei_col_idx = cols.index('IMEI') if 'IMEI' in cols else None

    ws.set_row(0, 30)
    ws.freeze_panes(1, 0)

    for c_idx, col_name in enumerate(cols):
        ws.set_column(c_idx, c_idx, col_widths.get(col_name, 15))
        ws.write(0, c_idx, col_name, hdr_fmt)

    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        for c_idx, val in enumerate(row):
            if c_idx == imei_col_idx:
                ws.write_string(r_idx, c_idx, str(val), str_fmt)
            else:
                ws.write(r_idx, c_idx, val, cell_fmt)

    wb.close()
    output.seek(0)
    return output


# ── Shared IMEI helpers (used by analytics, diagnostics, and daily delta) ─────
def _is_real_device(imei_val):
    """True only when IMEI is a genuine device — not blank, 0, or 'Empty pocket'."""
    s = str(imei_val).strip().lower()
    return s not in ('', 'nan', '0', 'none', 'empty pocket')

def _clean_imei(val):
    if pd.isna(val): return ''
    return str(val).strip().split('.')[0]


# ── Analytics Dashboard ───────────────────────────────────────────────────────
def show_analytics(hanger_file, totes_file, master_df):  # noqa: C901
    import plotly.express as px
    import plotly.graph_objects as go

    POCKETS = list('ABCDEFGHIJKLMNOPQRSTUVWX')

    # ── Dark dashboard CSS ────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .dash-section {
        background: linear-gradient(135deg, #0f0f1a 0%, #1a1a2e 100%);
        border-radius: 20px;
        padding: 2rem;
        margin: 1rem 0;
        border: 1px solid rgba(255,255,255,0.08);
    }
    .dash-title {
        font-size: 1.6rem;
        font-weight: 700;
        color: #ffffff;
        letter-spacing: 1px;
        margin-bottom: 0.3rem;
    }
    .dash-subtitle {
        font-size: 0.9rem;
        color: #aaaacc;
        margin-bottom: 1.5rem;
    }
    .hero-card {
        background: linear-gradient(135deg, rgba(255,255,255,0.05) 0%, rgba(255,255,255,0.02) 100%);
        border: 1px solid rgba(255,255,255,0.1);
        border-radius: 16px;
        padding: 1.5rem;
        text-align: center;
        backdrop-filter: blur(10px);
    }
    .hero-num {
        font-size: 3rem;
        font-weight: 800;
        line-height: 1;
        background: linear-gradient(90deg, #4472C4, #70AD47);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .hero-lbl {
        font-size: 0.85rem;
        color: #aaaacc;
        margin-top: 0.4rem;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    .hero-sub {
        font-size: 1rem;
        font-weight: 600;
        margin-top: 0.2rem;
    }
    .slot-grid {
        display: flex;
        flex-wrap: wrap;
        gap: 6px;
        margin-top: 0.5rem;
    }
    .slot-filled {
        width: 28px; height: 28px;
        border-radius: 50%;
        background: linear-gradient(135deg, #4472C4, #70AD47);
        display: flex; align-items: center; justify-content: center;
        font-size: 0.55rem; color: white; font-weight: 700;
        box-shadow: 0 0 8px rgba(68,114,196,0.6);
        cursor: default;
    }
    .slot-empty {
        width: 28px; height: 28px;
        border-radius: 50%;
        background: rgba(255,255,255,0.06);
        border: 1px dashed rgba(255,255,255,0.2);
        display: flex; align-items: center; justify-content: center;
        font-size: 0.55rem; color: rgba(255,255,255,0.3); font-weight: 700;
    }
    .hanger-label {
        font-size: 0.75rem;
        color: #aaaacc;
        font-weight: 600;
        margin-bottom: 4px;
        letter-spacing: 1px;
    }
    .hanger-block {
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 12px;
        padding: 0.75rem;
        margin-bottom: 0.75rem;
    }
    .section-divider {
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.15), transparent);
        margin: 2rem 0;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Load raw data ─────────────────────────────────────────────────────────
    hanger_raw = None
    totes_raw  = None

    if hanger_file:
        hanger_file.seek(0)
        h = pd.read_excel(hanger_file, sheet_name='Sheet1', engine='openpyxl')
        hanger_raw = h[h['Room'].notna()].reset_index(drop=True)

    if totes_file:
        totes_file.seek(0)
        t = pd.read_excel(totes_file, sheet_name='Sheet1', engine='openpyxl')
        totes_raw = t[t['Room'].notna()].reset_index(drop=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — HERO KPIs
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="dash-section">', unsafe_allow_html=True)
    st.markdown('<div class="dash-title">⚡ Operations Command Center</div>', unsafe_allow_html=True)
    st.markdown('<div class="dash-subtitle">Real-time overview of your warehouse and inventory status</div>', unsafe_allow_html=True)

    # compute top-level numbers
    hanger_raw = None
    if hanger_file:
        hanger_file.seek(0)
        h = pd.read_excel(hanger_file, sheet_name='Sheet1', engine='openpyxl')
        hanger_raw = h[h['Room'].notna()].reset_index(drop=True)

    totes_raw = None
    if totes_file:
        totes_file.seek(0)
        t = pd.read_excel(totes_file, sheet_name='Sheet1', engine='openpyxl')
        totes_raw = t[t['Room'].notna()].reset_index(drop=True)

    # Use master_df row count as ground truth when available — it covers ALL
    # uploaded hanger/totes files. Falling back to re-read counts only when
    # master_df hasn't been generated yet.
    if master_df is not None:
        total_devices = len(master_df)
    else:
        total_devices = (len(hanger_raw) if hanger_raw is not None else 0) + \
                        (len(totes_raw)  if totes_raw  is not None else 0)
    tracked        = len(master_df[master_df['Deal Id'] != 'No deal ID']) if master_df is not None else 0
    untracked      = total_devices - tracked
    track_pct      = round(tracked / total_devices * 100) if total_devices else 0

    occupied_slots = 0
    empty_slots    = 0
    hangers_int    = []
    occupied       = {}
    filled_counts  = {}
    empty_counts   = {}

    if hanger_raw is not None:
        hangers = sorted(hanger_raw['Hangar'].dropna().unique())
        hangers_int = [int(x) for x in hangers]
        for _, row in hanger_raw.iterrows():
            if pd.notna(row['Hangar']) and pd.notna(row['LOA']):
                if not _is_real_device(row.get('IMEI', '')):
                    continue          # skip rows that are physically empty pockets
                h_num = int(row['Hangar'])
                pocket = str(row['LOA']).strip().upper()
                occupied.setdefault(h_num, set()).add(pocket)
        filled_counts = {h: len(occupied.get(h, set())) for h in hangers_int}
        empty_counts  = {h: len([p for p in POCKETS if p not in occupied.get(h, set())]) for h in hangers_int}
        occupied_slots = sum(filled_counts.values())
        empty_slots    = sum(empty_counts.values())
        total_pockets  = len(POCKETS)
        fill_pct       = round(occupied_slots / (len(hangers_int) * total_pockets) * 100, 1) if hangers_int else 0
    else:
        total_pockets = len(POCKETS)
        fill_pct = 0

    h1, h2, h3, h4, h5 = st.columns(5)
    hero_cards = [
        (h1, str(total_devices),    "Total Devices",       f"{total_devices:,} devices in system",          "linear-gradient(135deg,#4472C4,#764ba2)"),
        (h2, f"{track_pct}%",       "Tracking Rate",       f"{tracked:,} fully tracked & identified",        "linear-gradient(135deg,#70AD47,#11998e)"),
        (h3, str(untracked),        "Need Attention",      "Devices missing deal ID",                        "linear-gradient(135deg,#f5576c,#f093fb)"),
        (h4, str(occupied_slots),   "Occupied Slots",      f"{fill_pct}% full · across {len(hangers_int)} hangers",   "linear-gradient(135deg,#4472C4,#3494E6)"),
        (h5, str(empty_slots),      "Available Slots",     f"{round(100 - fill_pct, 1)}% free · ready for new stock",  "linear-gradient(135deg,#FFC000,#f5576c)"),
    ]
    for col, num, lbl, sub, grad in hero_cards:
        with col:
            st.markdown(f"""
            <div class="hero-card" style="border-top: 3px solid transparent; border-image: {grad} 1;">
                <div class="hero-num">{num}</div>
                <div class="hero-lbl">{lbl}</div>
                <div class="hero-sub" style="color:#aaaacc;">{sub}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — OPERATIONS ACTION BOARD
    # ══════════════════════════════════════════════════════════════════════════
    act_col1, act_col2 = st.columns(2)

    # ── LEFT: Duplicate IMEI Detector ────────────────────────────────────────
    with act_col1:
        st.markdown('<div class="dash-section">', unsafe_allow_html=True)
        st.markdown('<div class="dash-title">🔁 Duplicate IMEI Alert</div>', unsafe_allow_html=True)
        st.markdown('<div class="dash-subtitle">Same IMEI found in more than one location — needs immediate resolution</div>', unsafe_allow_html=True)

        imei_locs: dict = {}
        if hanger_raw is not None:
            for _, row in hanger_raw.iterrows():
                imei = _clean_imei(row.get('IMEI', ''))
                if not _is_real_device(imei):
                    continue
                h_num = str(int(row['Hangar'])) if pd.notna(row.get('Hangar')) else '?'
                loa   = str(row.get('LOA', '')).strip()
                loc   = f"Hanger {h_num}{loa}"
                imei_locs.setdefault(imei, []).append(loc)
        if totes_raw is not None:
            for _, row in totes_raw.iterrows():
                imei = _clean_imei(row.get('IMEI', ''))
                if not _is_real_device(imei):
                    continue
                loa = str(row.get('LOA', '')).strip()
                imei_locs.setdefault(imei, []).append(f"Totes-{loa}")

        dup_rows = [
            {'IMEI': imei, 'Location 1': locs[0], 'Location 2': locs[1],
             'Extra Locations': ', '.join(locs[2:]) if len(locs) > 2 else '—'}
            for imei, locs in imei_locs.items() if len(locs) > 1
        ]

        if dup_rows:
            dup_df = pd.DataFrame(dup_rows)
            badge_color = '#E74C3C'
            badge_text  = f"{len(dup_df)} duplicate{'s' if len(dup_df) > 1 else ''} found"
        else:
            dup_df = pd.DataFrame(columns=['IMEI', 'Location 1', 'Location 2', 'Extra Locations'])
            badge_color = '#70AD47'
            badge_text  = 'No duplicates — all clear'

        st.markdown(f"""
        <div style="display:inline-block;background:{badge_color}22;color:{badge_color};
                    border:1px solid {badge_color}55;border-radius:8px;
                    padding:0.3rem 0.9rem;font-weight:700;font-size:0.9rem;margin-bottom:1rem;">
            {badge_text}
        </div>""", unsafe_allow_html=True)

        if dup_rows:
            st.dataframe(dup_df, use_container_width=True, hide_index=True, height=280)
            st.caption("⚠️ Use the button below — do NOT use the table's built-in export icon")
            dl_dup = export_excel(dup_df)
            st.download_button(
                label=f"⬇️ Download Duplicate IMEI List ({len(dup_df)} rows)",
                data=dl_dup,
                file_name="Duplicate_IMEIs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key='dl_dup_imei', type='primary', use_container_width=True,
            )
        else:
            st.success("Nothing to action here — no duplicate IMEIs detected across hangers and totes.")

        st.markdown('</div>', unsafe_allow_html=True)

    # ── RIGHT: Missing Data Fix List ──────────────────────────────────────────
    with act_col2:
        st.markdown('<div class="dash-section">', unsafe_allow_html=True)
        st.markdown('<div class="dash-title">🩹 Missing Data Fix List</div>', unsafe_allow_html=True)
        st.markdown('<div class="dash-subtitle">Devices with a Deal ID but incomplete info — fix these to improve tracking rate</div>', unsafe_allow_html=True)

        if master_df is not None:
            ae_mask = master_df['Deal Id'].str.contains('AE', na=False, case=False) & \
                      (master_df['Deal Id'] != 'No deal ID')
            ae_df = master_df[ae_mask].copy()

            missing_grade = ae_df['Grade'].isna() | (ae_df['Grade'].astype(str).str.strip() == '')
            missing_brand = ae_df['Brand'].isna() | (ae_df['Brand'].astype(str).str.strip() == '')
            missing_model = ae_df['Model'].isna() | (ae_df['Model'].astype(str).str.strip() == '')
            any_missing   = missing_grade | missing_brand | missing_model

            fix_df = ae_df[any_missing][['IMEI', 'Deal Id', 'Location', 'Brand', 'Model', 'Grade']].copy()
            fix_df['Missing Fields'] = ''
            fix_df.loc[missing_grade[any_missing], 'Missing Fields'] += 'Grade '
            fix_df.loc[missing_brand[any_missing], 'Missing Fields'] += 'Brand '
            fix_df.loc[missing_model[any_missing], 'Missing Fields'] += 'Model '
            fix_df['Missing Fields'] = fix_df['Missing Fields'].str.strip().str.replace(' ', ', ')
            fix_df = fix_df.reset_index(drop=True)

            n_fix = len(fix_df)
            n_grade = int(missing_grade.sum())
            n_brand = int(missing_brand.sum())
            n_model = int(missing_model.sum())
        else:
            fix_df = pd.DataFrame()
            n_fix = n_grade = n_brand = n_model = 0

        if n_fix > 0:
            badge_color2 = '#FFC000'
            badge_text2  = f"{n_fix} device{'s' if n_fix > 1 else ''} need fixing"
        else:
            badge_color2 = '#70AD47'
            badge_text2  = 'All tracked devices are complete'

        st.markdown(f"""
        <div style="display:inline-block;background:{badge_color2}22;color:{badge_color2};
                    border:1px solid {badge_color2}55;border-radius:8px;
                    padding:0.3rem 0.9rem;font-weight:700;font-size:0.9rem;margin-bottom:0.6rem;">
            {badge_text2}
        </div>""", unsafe_allow_html=True)

        if n_fix > 0:
            mc1, mc2, mc3 = st.columns(3)
            for col, val, lbl in [(mc1, n_grade, 'Missing Grade'),
                                   (mc2, n_brand, 'Missing Brand'),
                                   (mc3, n_model, 'Missing Model')]:
                with col:
                    st.markdown(f"""
                    <div style="background:rgba(255,192,0,0.08);border:1px solid rgba(255,192,0,0.2);
                                border-radius:10px;padding:0.6rem;text-align:center;">
                        <div style="font-size:1.6rem;font-weight:800;color:#FFC000;">{val}</div>
                        <div style="font-size:0.75rem;color:#aaaacc;">{lbl}</div>
                    </div>""", unsafe_allow_html=True)
            st.markdown('<br>', unsafe_allow_html=True)
            st.dataframe(fix_df, use_container_width=True, hide_index=True, height=220)
            st.caption("⚠️ Use the button below — do NOT use the table's built-in export icon")
            dl_fix = export_excel(fix_df)
            st.download_button(
                label=f"⬇️ Download Fix List ({n_fix} rows)",
                data=dl_fix,
                file_name="Missing_Data_Fix_List.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key='dl_fix_list', type='primary', use_container_width=True,
            )
        else:
            st.success("All devices with Deal IDs have complete Brand, Model and Grade data.")

        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3 — HANGER SLOT MAP (physical dot grid)
    # ══════════════════════════════════════════════════════════════════════════
    if hanger_raw is not None:
        st.markdown('<div class="dash-section">', unsafe_allow_html=True)
        st.markdown('<div class="dash-title">🗄️ Available Slots Map</div>', unsafe_allow_html=True)
        st.markdown('<div class="dash-subtitle">Only showing empty pockets — these are ready for new devices to be stored</div>', unsafe_allow_html=True)

        # Only show hangers that have at least 1 empty pocket
        hangers_with_space = [h for h in hangers_int if empty_counts.get(h, 0) > 0]

        if not hangers_with_space:
            st.markdown('<div style="text-align:center;color:#70AD47;font-size:1.2rem;padding:2rem;">🎉 All pockets are filled across every hanger!</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="color:#aaaacc;margin-bottom:1rem;">Showing <b style="color:white;">{len(hangers_with_space)}</b> hangers with available space &nbsp;·&nbsp; <b style="color:#555;">{len(hangers_int) - len(hangers_with_space)}</b> hangers fully occupied (hidden)</div>', unsafe_allow_html=True)

            cols_per_row = 5
            hanger_chunks = [hangers_with_space[i:i+cols_per_row] for i in range(0, len(hangers_with_space), cols_per_row)]

            for chunk in hanger_chunks:
                cols = st.columns(len(chunk))
                for col, h_num in zip(cols, chunk):
                    with col:
                        empty_pockets = [p for p in POCKETS if p not in occupied.get(h_num, set())]
                        empty_count   = len(empty_pockets)
                        color = '#FFC000' if empty_count <= 6 else '#4472C4'

                        pockets_html = ''.join(
                            f'<div class="slot-empty" style="border-color:{color};color:{color};" title="Hanger {h_num} — Pocket {p} — Available">{p}</div>'
                            for p in empty_pockets
                        )

                        st.markdown(f"""
                        <div class="hanger-block">
                            <div class="hanger-label">
                                HANGER {h_num} &nbsp;·&nbsp;
                                <span style="color:{color};">{empty_count} EMPTY</span>
                            </div>
                            <div class="slot-grid">{pockets_html}</div>
                        </div>
                        """, unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4 — INVENTORY COMPOSITION
    # ══════════════════════════════════════════════════════════════════════════
    if master_df is not None:
        matched = master_df[master_df['Category'] != '']
        if len(matched) > 0:
            st.markdown('<div class="dash-section">', unsafe_allow_html=True)
            st.markdown('<div class="dash-title">📦 What\'s In Your Warehouse</div>', unsafe_allow_html=True)
            st.markdown('<div class="dash-subtitle">Breakdown of devices by brand, grade and origin country</div>', unsafe_allow_html=True)

            # Treemap — brands & models
            brand_model = matched.groupby(['Brand', 'Model']).size().reset_index(name='Count')
            fig_tree = px.treemap(
                brand_model, path=['Brand', 'Model'], values='Count',
                color='Count',
                color_continuous_scale=['#1a1a2e', '#4472C4', '#70AD47'],
            )
            fig_tree.update_layout(
                height=420,
                margin=dict(l=0, r=0, t=0, b=0),
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
            )
            fig_tree.update_traces(
                textfont=dict(color='white', size=13),
                hovertemplate='<b>%{label}</b><br>%{value} devices<extra></extra>',
            )
            st.plotly_chart(fig_tree, use_container_width=True)

            # ── Drill-down: Brand → Model → Download ─────────────────────────
            st.markdown('<div style="height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.12),transparent);margin:1.5rem 0;"></div>', unsafe_allow_html=True)
            st.markdown('<div style="color:white;font-weight:700;font-size:1.05rem;margin-bottom:0.6rem;">🔎 Drill Down & Export — click a box above then filter here</div>', unsafe_allow_html=True)

            dd1, dd2, dd3 = st.columns([2, 3, 2])

            with dd1:
                brand_options = sorted(matched['Brand'].dropna().unique().tolist())
                selected_brand = st.selectbox(
                    "Select Brand",
                    options=brand_options,
                    index=None,
                    placeholder="Type to search brands...",
                    key='drill_brand',
                )

            filtered_by_brand = matched if selected_brand is None else matched[matched['Brand'] == selected_brand]

            with dd2:
                model_options = sorted(filtered_by_brand['Model'].dropna().unique().tolist())
                selected_model = st.selectbox(
                    "Select Model",
                    options=model_options,
                    index=None,
                    placeholder="Type to search models...",
                    key='drill_model',
                )

            drill_df = filtered_by_brand if selected_model is None else filtered_by_brand[filtered_by_brand['Model'] == selected_model]

            with dd3:
                st.markdown('<div style="padding-top:1.8rem;">', unsafe_allow_html=True)
                drill_label = selected_brand if selected_brand is not None else "All Brands"
                if selected_model is not None:
                    drill_label += f" · {selected_model}"
                st.markdown(f'<div style="color:#aaaacc;font-size:0.82rem;margin-bottom:0.4rem;">{len(drill_df):,} devices matched</div>', unsafe_allow_html=True)
                if len(drill_df) > 0:
                    drill_excel = export_excel(drill_df.reset_index(drop=True))
                    safe_label  = drill_label.replace(' ', '_').replace('·', '').replace('/', '-')[:60]
                    st.download_button(
                        label=f"⬇️ Download ({len(drill_df):,} rows)",
                        data=drill_excel,
                        file_name=f"Export_{safe_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key='drill_download',
                    )
                st.markdown('</div>', unsafe_allow_html=True)

            if len(drill_df) > 0:
                st.dataframe(drill_df.reset_index(drop=True), use_container_width=True, hide_index=True)
                st.caption(f"Showing all {len(drill_df):,} rows for: {drill_label}")

            c1, c2 = st.columns(2)

            # Build a dynamic grade colour palette from the actual grades in the data
            GRADE_PALETTE = [
                '#11998e','#70AD47','#4472C4','#FFC000','#E74C3C',
                '#8E44AD','#EC6EAD','#3494E6','#f5576c','#f093fb',
                '#38ef7d','#FF6B35','#A8E6CF','#FFD93D','#6BCB77',
            ]
            all_grades = matched['Grade'].dropna().unique().tolist()
            grade_color_map = {g: GRADE_PALETTE[i % len(GRADE_PALETTE)] for i, g in enumerate(sorted(all_grades))}

            with c1:
                grade_counts = matched['Grade'].value_counts().reset_index()
                grade_counts.columns = ['Grade', 'Count']
                pie_colors = [grade_color_map.get(g, '#4472C4') for g in grade_counts['Grade']]
                fig_grade = go.Figure(go.Pie(
                    labels=grade_counts['Grade'],
                    values=grade_counts['Count'],
                    hole=0.50,
                    marker=dict(colors=pie_colors,
                                line=dict(color='rgba(0,0,0,0)', width=0)),
                    textfont=dict(color='white', size=12),
                    hovertemplate='%{label}<br>%{value} devices (%{percent})<extra></extra>',
                ))
                fig_grade.update_layout(
                    title=dict(text='Grade Breakdown', font=dict(color='white', size=15), x=0.5),
                    height=400,
                    margin=dict(l=10, r=10, t=50, b=10),
                    paper_bgcolor='rgba(0,0,0,0)',
                    legend=dict(font=dict(color='white', size=11)),
                    annotations=[dict(text=f"<b>{len(matched)}</b><br>devices", x=0.5, y=0.5,
                                      font=dict(size=13, color='white'), showarrow=False)],
                )
                st.plotly_chart(fig_grade, use_container_width=True)

                # ── Grade download ────────────────────────────────────────────
                grade_options = sorted(matched['Grade'].dropna().unique().tolist())
                sel_grade = st.selectbox(
                    "Download devices by grade",
                    options=grade_options,
                    index=None,
                    placeholder="Select a grade to download...",
                    key='grade_dl_select',
                )
                if sel_grade is not None:
                    grade_df = matched[matched['Grade'] == sel_grade].reset_index(drop=True)
                    st.caption(f"{len(grade_df):,} devices with grade **{sel_grade}**")
                    # Cache bytes in session state — only recompute when grade changes
                    cache_key = f'_grade_dl_bytes_{sel_grade}'
                    if cache_key not in st.session_state:
                        with st.spinner(f"Preparing {sel_grade} export..."):
                            st.session_state[cache_key] = export_excel(grade_df).read()
                    st.download_button(
                        label=f"⬇️ Download {sel_grade} devices ({len(grade_df):,} rows)",
                        data=st.session_state[cache_key],
                        file_name=f"Grade_{sel_grade.replace(' ', '_')}_Devices.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key='dl_grade_export',
                        type='primary',
                        use_container_width=True,
                    )

            with c2:
                vat_counts = matched['VAT Type'].value_counts().reset_index()
                vat_counts.columns = ['VAT Type', 'Count']
                fig_vat = go.Figure(go.Pie(
                    labels=vat_counts['VAT Type'],
                    values=vat_counts['Count'],
                    hole=0.50,
                    marker=dict(colors=['#4472C4','#70AD47','#FFC000','#E74C3C'],
                                line=dict(color='rgba(0,0,0,0)', width=0)),
                    textfont=dict(color='white', size=12),
                    hovertemplate='%{label}<br>%{value} devices (%{percent})<extra></extra>',
                ))
                fig_vat.update_layout(
                    title=dict(text='VAT Type', font=dict(color='white', size=15), x=0.5),
                    height=400,
                    margin=dict(l=10, r=10, t=50, b=10),
                    paper_bgcolor='rgba(0,0,0,0)',
                    legend=dict(font=dict(color='white', size=11)),
                    annotations=[dict(text=f"<b>{len(matched)}</b><br>devices", x=0.5, y=0.5,
                                      font=dict(size=13, color='white'), showarrow=False)],
                )
                st.plotly_chart(fig_vat, use_container_width=True)

            st.markdown('</div>', unsafe_allow_html=True)

        # ── Hanger vs Totes ───────────────────────────────────────────────────
        if 'Bin' in master_df.columns:
            bin_counts = master_df['Bin'].value_counts().reset_index()
            bin_counts.columns = ['Type', 'Count']
            if len(bin_counts) > 1:
                st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
                st.markdown('<div class="dash-section">', unsafe_allow_html=True)
                st.markdown('<div class="dash-title">⚖️ Hanger vs Totes Split</div>', unsafe_allow_html=True)
                st.markdown('<div class="dash-subtitle">Distribution of devices across storage types</div>', unsafe_allow_html=True)
                fig_split = go.Figure(go.Pie(
                    labels=bin_counts['Type'],
                    values=bin_counts['Count'],
                    hole=0.6,
                    marker=dict(colors=['#4472C4','#70AD47'],
                                line=dict(color='rgba(0,0,0,0)', width=0)),
                    textfont=dict(color='white', size=14),
                    hovertemplate='%{label}<br>%{value} devices (%{percent})<extra></extra>',
                ))
                fig_split.update_layout(
                    height=320,
                    margin=dict(l=20, r=20, t=20, b=20),
                    paper_bgcolor='rgba(0,0,0,0)',
                    legend=dict(font=dict(color='white', size=13)),
                )
                st.plotly_chart(fig_split, use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 5 — FULL INTERACTIVE HANGER MAP WITH DEVICE DETAILS ON HOVER
    # ══════════════════════════════════════════════════════════════════════════
    if hanger_raw is not None:
        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="dash-section">', unsafe_allow_html=True)
        st.markdown('<div class="dash-title">🔍 Interactive Hanger Map</div>', unsafe_allow_html=True)
        st.markdown('<div class="dash-subtitle">Hover over any pocket to see what\'s stored — every hanger, every slot</div>', unsafe_allow_html=True)

        # Normalize IMEI same way as process_hanger so lookup matches master_df
        def _norm_imei(val):
            if pd.isna(val):
                return '—'
            if isinstance(val, float):
                return str(int(val))
            return str(val).strip()

        # Build master_df IMEI index for fast lookup
        master_imei_index = {}
        if master_df is not None:
            for _, mrow in master_df.iterrows():
                key = str(mrow.get('IMEI', '')).strip()
                if key and key != 'nan':
                    master_imei_index[key] = mrow

        # Build device lookup: { (hanger_num, pocket) → device details }
        # Only add entry if the row has a real IMEI (not "Empty pocket" / 0 / blank)
        device_lookup = {}
        for _, row in hanger_raw.iterrows():
            if pd.notna(row['Hangar']) and pd.notna(row['LOA']):
                if not _is_real_device(row.get('IMEI', '')):
                    continue          # genuine empty pocket — leave out of device_lookup
                h_num  = int(row['Hangar'])
                pocket = str(row['LOA']).strip().upper()
                imei   = _norm_imei(row['IMEI'])

                brand = model = grade = deal_id = status = '—'
                if imei != '—' and imei in master_imei_index:
                    m = master_imei_index[imei]
                    brand   = str(m.get('Brand',   '') or '').strip() or '—'
                    model   = str(m.get('Model',   '') or '').strip() or '—'
                    grade   = str(m.get('Grade',   '') or '').strip() or '—'
                    deal_id = str(m.get('Deal Id', '') or '').strip() or '—'
                    status  = str(m.get('Status',  '') or '').strip() or '—'

                device_lookup[(h_num, pocket)] = {
                    'imei':    imei,
                    'brand':   brand,
                    'model':   model,
                    'grade':   grade,
                    'deal_id': deal_id,
                    'status':  status,
                }

        # Build scatter data points
        x_vals, y_vals, colors, hover_texts, sizes = [], [], [], [], []

        for h_num in hangers_int:
            for pocket in POCKETS:
                x_vals.append(str(h_num))
                y_vals.append(pocket)

                if pocket in occupied.get(h_num, set()):
                    dev = device_lookup.get((h_num, pocket), {})
                    brand   = dev.get('brand',   '—')
                    model   = dev.get('model',   '—')
                    grade   = dev.get('grade',   '—')
                    deal_id = dev.get('deal_id', '—')
                    imei    = dev.get('imei',    '—')
                    status  = dev.get('status',  '—')

                    # Color by grade
                    grade_colors = {
                        'Grade A':    '#70AD47',
                        'Grade A+':   '#11998e',
                        'Grade B':    '#4472C4',
                        'Grade C':    '#FFC000',
                        'Grade D':    '#E74C3C',
                    }
                    dot_color = grade_colors.get(grade, '#4472C4')
                    colors.append(dot_color)
                    sizes.append(14)

                    hover_texts.append(
                        f"<b>📍 Hanger {h_num} — Pocket {pocket}</b><br>"
                        f"──────────────────<br>"
                        f"<b>Brand:</b> {brand}<br>"
                        f"<b>Model:</b> {model}<br>"
                        f"<b>Grade:</b> {grade}<br>"
                        f"<b>Deal ID:</b> {deal_id}<br>"
                        f"<b>IMEI:</b> {imei}<br>"
                        f"<b>Status:</b> {status}"
                    )
                else:
                    colors.append('rgba(255,255,255,0.08)')
                    sizes.append(10)
                    hover_texts.append(
                        f"<b>📍 Hanger {h_num} — Pocket {pocket}</b><br>"
                        f"──────────────────<br>"
                        f"<i>Empty Pocket — Available for new stock</i>"
                    )

        # Build dynamic grade colour map from all grades present in device_lookup
        MAP_PALETTE = [
            '#11998e','#70AD47','#FFC000','#E74C3C','#8E44AD',
            '#3494E6','#EC6EAD','#f5576c','#38ef7d','#FF6B35',
            '#A8E6CF','#FFD93D','#6BCB77','#4472C4','#f093fb',
        ]
        all_present_grades = sorted({
            v['grade'] for v in device_lookup.values() if v['grade'] != '—'
        })
        grade_colors = {g: MAP_PALETTE[i % len(MAP_PALETTE)] for i, g in enumerate(all_present_grades)}

        # Legend — built dynamically from whatever grades exist in the data
        legend_items = ''.join(
            f'<span style="color:{grade_colors[g]};">● {g}</span>'
            for g in all_present_grades
        )
        legend_items += '<span style="color:rgba(255,255,255,0.2);">● Empty Pocket</span>'
        st.markdown(
            f'<div style="display:flex;gap:1.5rem;margin-bottom:1.2rem;flex-wrap:wrap;font-size:0.9rem;">'
            f'{legend_items}</div>',
            unsafe_allow_html=True
        )

        # Each hanger: 3 pockets per row (A B C / D E F / ... / V W X) = 8 rows × 3 cols
        COLS_PER_ROW = 3
        # Build grid positions for 24 pockets
        pocket_grid = {}  # pocket → (col_x, row_y)
        for idx, p in enumerate(POCKETS):
            pocket_grid[p] = (idx % COLS_PER_ROW, idx // COLS_PER_ROW)

        # ── Range selector ────────────────────────────────────────────────────
        total_hangers = len(hangers_int)

        # Initialise stored range (separate from widget key)
        _default_range = (1, min(8, total_hangers))
        if 'hmap_range' not in st.session_state:
            st.session_state['hmap_range'] = _default_range

        sl_col, toggle_col = st.columns([5, 1])
        with toggle_col:
            manual_mode = st.checkbox("✏️ Manual", key='hmap_manual_mode',
                                      help="Check to type exact hanger numbers")

        if manual_mode:
            with sl_col:
                mc1, mc2 = st.columns(2)
                with mc1:
                    range_start = st.number_input("From hanger", min_value=1, max_value=total_hangers,
                                                   value=st.session_state['hmap_range'][0],
                                                   step=1, key='hmap_manual_start')
                with mc2:
                    range_end = st.number_input("To hanger", min_value=1, max_value=total_hangers,
                                                 value=st.session_state['hmap_range'][1],
                                                 step=1, key='hmap_manual_end')
            range_start = max(1, min(int(range_start), total_hangers))
            range_end   = max(range_start, min(int(range_end), total_hangers))
            st.session_state['hmap_range'] = (range_start, range_end)
        else:
            with sl_col:
                range_start, range_end = st.slider(
                    "Hanger range",
                    min_value=1, max_value=total_hangers,
                    value=st.session_state['hmap_range'],
                    step=1, key='hmap_slider_widget',
                )
            st.session_state['hmap_range'] = (range_start, range_end)

        visible_hangers = hangers_int[range_start - 1 : range_end]
        st.markdown(
            f'<div style="color:#aaaacc;font-size:0.82rem;margin-bottom:0.8rem;">'
            f'Showing hangers <b style="color:white;">{hangers_int[range_start-1]}</b> → '
            f'<b style="color:white;">{hangers_int[range_end-1]}</b> '
            f'&nbsp;·&nbsp; {len(visible_hangers)} of {total_hangers} hangers</div>',
            unsafe_allow_html=True
        )

        # 4 hangers per row across the page — always 4 columns so width stays consistent
        hangers_per_row = 4
        hanger_chunks   = [visible_hangers[i:i+hangers_per_row] for i in range(0, len(visible_hangers), hangers_per_row)]

        for chunk in hanger_chunks:
            cols = st.columns(hangers_per_row)
            for col, h_num in zip(cols, chunk):
                dot_x, dot_y, dot_colors, dot_sizes, dot_text, dot_hover = [], [], [], [], [], []

                for pocket in POCKETS:
                    gx, gy = pocket_grid[pocket]
                    dot_x.append(gx)
                    dot_y.append(gy)
                    dot_text.append(pocket)

                    if pocket in occupied.get(h_num, set()):
                        dev     = device_lookup.get((h_num, pocket), {})
                        brand   = dev.get('brand',   '—')
                        model   = dev.get('model',   '—')
                        grade   = dev.get('grade',   '—')
                        deal_id = dev.get('deal_id', '—')
                        imei    = dev.get('imei',    '—')
                        status  = dev.get('status',  '—')
                        dot_colors.append(grade_colors.get(grade, '#4472C4'))
                        dot_sizes.append(28)
                        dot_hover.append(
                            f"<b>Hanger {h_num} · Pocket {pocket}</b><br>"
                            f"─────────────────<br>"
                            f"<b>Brand:</b> {brand}<br>"
                            f"<b>Model:</b> {model}<br>"
                            f"<b>Grade:</b> {grade}<br>"
                            f"<b>Deal ID:</b> {deal_id}<br>"
                            f"<b>IMEI:</b> {imei}<br>"
                            f"<b>Status:</b> {status}"
                        )
                    else:
                        dot_colors.append('rgba(255,255,255,0.07)')
                        dot_sizes.append(24)
                        dot_hover.append(
                            f"<b>Hanger {h_num} · Pocket {pocket}</b><br>"
                            f"─────────────────<br>"
                            f"<i>Empty — Available for new stock</i>"
                        )

                filled  = filled_counts.get(h_num, 0)
                empty_n = empty_counts.get(h_num, 0)
                rate    = round(filled / total_pockets * 100)
                rate_color = '#70AD47' if rate >= 70 else ('#FFC000' if rate >= 40 else '#E74C3C')

                fig_h = go.Figure(go.Scatter(
                    x=dot_x,
                    y=dot_y,
                    mode='markers+text',
                    text=dot_text,
                    textposition='middle center',
                    textfont=dict(color='white', size=8, family='Calibri'),
                    marker=dict(
                        size=dot_sizes,
                        color=dot_colors,
                        line=dict(color='rgba(255,255,255,0.12)', width=1),
                        symbol='circle',
                    ),
                    hovertemplate='%{customdata}<extra></extra>',
                    customdata=dot_hover,
                ))

                fig_h.update_layout(
                    title=dict(
                        text=f'<b style="color:white;">Hanger {h_num}</b>  '
                             f'<span style="color:{rate_color};font-size:11px;">{rate}% occupied · {filled} filled · {100 - rate}% empty · {empty_n} empty</span>',
                        font=dict(size=12, color='white'),
                        x=0,
                    ),
                    height=280,
                    margin=dict(l=10, r=10, t=40, b=10),
                    paper_bgcolor='rgba(255,255,255,0.03)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(
                        showticklabels=False,
                        showgrid=False,
                        zeroline=False,
                        fixedrange=True,
                        range=[-0.5, 2.5],
                    ),
                    yaxis=dict(
                        showticklabels=False,
                        showgrid=False,
                        zeroline=False,
                        fixedrange=True,
                        autorange='reversed',
                        range=[-0.5, 7.5],
                    ),
                    hoverlabel=dict(
                        bgcolor='#1a1a2e',
                        bordercolor='rgba(68,114,196,0.6)',
                        font=dict(color='white', size=12, family='Calibri'),
                    ),
                )

                with col:
                    st.plotly_chart(fig_h, use_container_width=True, config={'displayModeBar': False})

        st.markdown('</div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 6 — LOW VALUE INVENTORY ANALYTICS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="dash-section">', unsafe_allow_html=True)
    st.markdown('<div class="dash-title">🏷️ Low Value Inventory Analytics</div>', unsafe_allow_html=True)
    st.markdown('<div class="dash-subtitle">Breakdown of low value products — Room = Inventory, Bin = Totes</div>', unsafe_allow_html=True)

    lv_df = None
    if master_df is not None:
        lv_df = master_df[master_df['Room'].astype(str).str.strip() == 'Inventory'].copy()
        lv_df = lv_df[lv_df['IMEI'].astype(str).str.strip() != ''].reset_index(drop=True)

    if lv_df is None or len(lv_df) == 0:
        st.info("No Low Value products in the current master template. Upload a Low Value file via the Totes button and generate to see analytics here.")
    else:
        # ── KPI row ───────────────────────────────────────────────────────────
        lv_total      = len(lv_df)
        lv_with_deal  = int((lv_df['Deal Id'].astype(str).str.strip().str.upper().str.startswith('AE')).sum())
        lv_brands     = lv_df['Brand'].astype(str).str.strip().replace('', pd.NA).dropna().nunique()
        lv_categories = lv_df['Category'].astype(str).str.strip().replace('', pd.NA).dropna().nunique()

        lv1, lv2, lv3, lv4 = st.columns(4)
        lv_kpis = [
            (lv1, str(lv_total),      "Low Value Devices",  "linear-gradient(135deg,#9B59B6,#764ba2)"),
            (lv2, str(lv_with_deal),  "With AE Deal ID",    "linear-gradient(135deg,#70AD47,#11998e)"),
            (lv3, str(lv_brands),     "Unique Brands",      "linear-gradient(135deg,#4472C4,#3494E6)"),
            (lv4, str(lv_categories), "Categories",         "linear-gradient(135deg,#FFC000,#f5576c)"),
        ]
        for col, num, lbl, grad in lv_kpis:
            with col:
                st.markdown(f"""
                <div class="hero-card" style="border-top: 3px solid transparent; border-image: {grad} 1;">
                    <div class="hero-num" style="font-size:2.2rem;">{num}</div>
                    <div class="hero-lbl">{lbl}</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Row 1: Category bar + Brand bar ───────────────────────────────────
        chart_col1, chart_col2 = st.columns(2)

        with chart_col1:
            cat_counts = (
                lv_df['Category'].astype(str).str.strip()
                .replace('', 'Unknown').replace('nan', 'Unknown')
                .value_counts().reset_index()
            )
            cat_counts.columns = ['Category', 'Count']
            fig_cat = go.Figure(go.Bar(
                x=cat_counts['Count'],
                y=cat_counts['Category'],
                orientation='h',
                marker=dict(
                    color=cat_counts['Count'],
                    colorscale='Purples',
                    showscale=False,
                ),
                text=cat_counts['Count'],
                textposition='outside',
                textfont=dict(color='white', size=11),
            ))
            fig_cat.update_layout(
                title=dict(text='<b style="color:white;">Devices by Category</b>', font=dict(size=13, color='white')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                height=320,
                margin=dict(l=10, r=40, t=40, b=10),
                xaxis=dict(showgrid=False, color='#aaaacc', zeroline=False),
                yaxis=dict(color='#aaaacc', autorange='reversed'),
                font=dict(color='white'),
            )
            st.plotly_chart(fig_cat, use_container_width=True, config={'displayModeBar': False})

        with chart_col2:
            brand_counts = (
                lv_df['Brand'].astype(str).str.strip()
                .replace('', 'Unknown').replace('nan', 'Unknown')
                .value_counts().head(10).reset_index()
            )
            brand_counts.columns = ['Brand', 'Count']
            fig_brand = go.Figure(go.Bar(
                x=brand_counts['Count'],
                y=brand_counts['Brand'],
                orientation='h',
                marker=dict(
                    color=brand_counts['Count'],
                    colorscale='Blues',
                    showscale=False,
                ),
                text=brand_counts['Count'],
                textposition='outside',
                textfont=dict(color='white', size=11),
            ))
            fig_brand.update_layout(
                title=dict(text='<b style="color:white;">Top 10 Brands</b>', font=dict(size=13, color='white')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                height=320,
                margin=dict(l=10, r=40, t=40, b=10),
                xaxis=dict(showgrid=False, color='#aaaacc', zeroline=False),
                yaxis=dict(color='#aaaacc', autorange='reversed'),
                font=dict(color='white'),
            )
            st.plotly_chart(fig_brand, use_container_width=True, config={'displayModeBar': False})

        # ── Row 2: Location distribution + Status pie ─────────────────────────
        chart_col3, chart_col4 = st.columns(2)

        with chart_col3:
            loc_counts = (
                lv_df['Location'].astype(str).str.strip()
                .replace('', 'Unknown').replace('nan', 'Unknown')
                .value_counts().reset_index()
            )
            loc_counts.columns = ['Location', 'Count']
            fig_loc = go.Figure(go.Bar(
                x=loc_counts['Location'],
                y=loc_counts['Count'],
                marker=dict(
                    color=loc_counts['Count'],
                    colorscale='Teal',
                    showscale=False,
                ),
                text=loc_counts['Count'],
                textposition='outside',
                textfont=dict(color='white', size=11),
            ))
            fig_loc.update_layout(
                title=dict(text='<b style="color:white;">Devices by Bin / Location</b>', font=dict(size=13, color='white')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                height=320,
                margin=dict(l=10, r=10, t=40, b=40),
                xaxis=dict(color='#aaaacc', tickangle=-45),
                yaxis=dict(showgrid=False, color='#aaaacc', zeroline=False),
                font=dict(color='white'),
            )
            st.plotly_chart(fig_loc, use_container_width=True, config={'displayModeBar': False})

        with chart_col4:
            status_counts = (
                lv_df['Status'].astype(str).str.strip()
                .replace('', 'Unknown').replace('nan', 'Unknown')
                .value_counts().reset_index()
            )
            status_counts.columns = ['Status', 'Count']
            colors_status = ['#9B59B6', '#4472C4', '#70AD47', '#FFC000', '#E74C3C',
                             '#3498DB', '#2ECC71', '#E67E22', '#1ABC9C', '#E74C3C']
            fig_status = go.Figure(go.Pie(
                labels=status_counts['Status'],
                values=status_counts['Count'],
                hole=0.45,
                marker=dict(colors=colors_status[:len(status_counts)], line=dict(color='#0f0f1a', width=2)),
                textinfo='label+percent',
                textfont=dict(color='white', size=11),
                hovertemplate='<b>%{label}</b><br>%{value} devices<br>%{percent}<extra></extra>',
            ))
            fig_status.update_layout(
                title=dict(text='<b style="color:white;">Status Breakdown</b>', font=dict(size=13, color='white')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                height=320,
                margin=dict(l=10, r=10, t=40, b=10),
                legend=dict(font=dict(color='white', size=10), bgcolor='rgba(0,0,0,0)'),
                font=dict(color='white'),
            )
            st.plotly_chart(fig_status, use_container_width=True, config={'displayModeBar': False})

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SMART DIAGNOSTICS DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
def show_diagnostics(hanger_file, totes_file, master_df):  # noqa: C901
    import plotly.graph_objects as go
    import plotly.express as px
    import re

    POCKETS = list('ABCDEFGHIJKLMNOPQRSTUVWX')

    st.markdown("""
    <style>
    .diag-header {
        font-size: 2rem; font-weight: 800; color: #fff;
        background: linear-gradient(90deg,#f5576c,#f093fb);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        margin-bottom: 0.3rem;
    }
    .diag-section {
        background: linear-gradient(135deg,#0f0f1a 0%,#1a1a2e 100%);
        border-radius: 20px; padding: 2rem; margin: 1.2rem 0;
        border: 1px solid rgba(255,255,255,0.08);
    }
    .diag-title  { font-size:1.4rem; font-weight:700; color:#fff; margin-bottom:0.2rem; }
    .diag-sub    { font-size:0.88rem; color:#aaaacc; margin-bottom:1.2rem; }
    .badge-red   { background:#E74C3C22; color:#E74C3C; border:1px solid #E74C3C55;
                   border-radius:8px; padding:0.2rem 0.6rem; font-size:0.8rem; font-weight:700; }
    .badge-green { background:#70AD4722; color:#70AD47; border:1px solid #70AD4755;
                   border-radius:8px; padding:0.2rem 0.6rem; font-size:0.8rem; font-weight:700; }
    .badge-blue  { background:#4472C422; color:#4472C4; border:1px solid #4472C455;
                   border-radius:8px; padding:0.2rem 0.6rem; font-size:0.8rem; font-weight:700; }
    .badge-gold  { background:#FFC00022; color:#FFC000; border:1px solid #FFC00055;
                   border-radius:8px; padding:0.2rem 0.6rem; font-size:0.8rem; font-weight:700; }
    .diag-divider { height:1px; background:linear-gradient(90deg,transparent,rgba(255,255,255,0.12),transparent); margin:2rem 0; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="diag-header">🔬 Smart Diagnostics</div>', unsafe_allow_html=True)
    st.markdown('<p style="color:#aaaacc;margin-bottom:1.5rem;">Deep-dive analysis to catch issues, reduce manual effort and optimise your warehouse</p>', unsafe_allow_html=True)

    # ── helpers ───────────────────────────────────────────────────────────────
    # ── load raw data ─────────────────────────────────────────────────────────
    hanger_raw = totes_raw = None
    if hanger_file:
        hanger_file.seek(0)
        h = pd.read_excel(hanger_file, sheet_name='Sheet1', engine='openpyxl')
        hanger_raw = h[h['Room'].notna()].reset_index(drop=True)
    if totes_file:
        totes_file.seek(0)
        t = pd.read_excel(totes_file, sheet_name='Sheet1', engine='openpyxl')
        totes_raw = t[t['Room'].notna()].reset_index(drop=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 1 — DUPLICATE IMEI DETECTOR
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="diag-section">', unsafe_allow_html=True)
    st.markdown('<div class="diag-title">🔁 Duplicate IMEI Detector</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-sub">Same IMEI appearing in two different locations — data entry error that silently breaks lookups</div>', unsafe_allow_html=True)

    all_rows = []
    if hanger_raw is not None:
        for _, row in hanger_raw.iterrows():
            if not _is_real_device(row.get('IMEI', '')): continue
            imei = _clean_imei(row['IMEI'])
            loc  = (str(int(row['Hangar'])) if pd.notna(row.get('Hangar')) else '') + str(row.get('LOA', '')).strip()
            all_rows.append({'IMEI': imei, 'Source': 'Hanger', 'Location': loc,
                             'Deal ID': str(row.get('Deal ID', '')).strip(),
                             'BACKEND': str(row.get('BACKEND', '')).strip()})
    if totes_raw is not None:
        for _, row in totes_raw.iterrows():
            if not _is_real_device(row.get('IMEI', '')): continue
            imei = _clean_imei(row['IMEI'])
            loc  = str(row.get('LOA', '')).strip()
            all_rows.append({'IMEI': imei, 'Source': 'Totes', 'Location': loc,
                             'Deal ID': str(row.get('Deal ID', '')).strip(),
                             'BACKEND': str(row.get('BACKEND', '')).strip()})

    if all_rows:
        imei_df  = pd.DataFrame(all_rows)
        dup_df   = imei_df[imei_df.duplicated('IMEI', keep=False)].sort_values('IMEI').reset_index(drop=True)
        dup_count = dup_df['IMEI'].nunique()

        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f'<div style="font-size:2.5rem;font-weight:800;color:{"#E74C3C" if dup_count else "#70AD47"};">{dup_count}</div>'
                        f'<div style="color:#aaaacc;font-size:0.9rem;">Duplicate IMEIs found</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div style="font-size:2.5rem;font-weight:800;color:#4472C4;">{len(imei_df)}</div>'
                        f'<div style="color:#aaaacc;font-size:0.9rem;">Total unique devices scanned</div>', unsafe_allow_html=True)

        if dup_count:
            st.markdown(f'<br><span class="badge-red">⚠️ {len(dup_df)} rows flagged — {dup_count} IMEIs appear more than once</span>', unsafe_allow_html=True)
            st.dataframe(dup_df, use_container_width=True, hide_index=True)
        else:
            st.markdown('<br><span class="badge-green">✅ No duplicate IMEIs — data is clean</span>', unsafe_allow_html=True)
    else:
        st.info("Upload at least one stocktake file to scan for duplicates.")

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 2 — NO DEAL ID HEATMAP BY HANGER
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="diag-section">', unsafe_allow_html=True)
    st.markdown('<div class="diag-title">📍 Untracked Devices by Hanger</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-sub">Which hangers have the most devices with no Deal ID — tells your team exactly where to investigate first</div>', unsafe_allow_html=True)

    if master_df is not None and hanger_raw is not None:
        hanger_master = master_df[master_df['Bin'] == 'Hanger'].copy()
        # Extract hanger number from Location (e.g. "61A" → 61)
        hanger_master['Hanger#'] = hanger_master['Location'].str.extract(r'^(\d+)').astype(float)
        no_deal = hanger_master[hanger_master['Deal Id'] == 'No deal ID']
        total_by_h = hanger_master.groupby('Hanger#').size().reset_index(name='Total')
        no_deal_by_h = no_deal.groupby('Hanger#').size().reset_index(name='No Deal ID')
        merged = total_by_h.merge(no_deal_by_h, on='Hanger#', how='left').fillna(0)
        merged['No Deal ID'] = merged['No Deal ID'].astype(int)
        merged['% Untracked'] = (merged['No Deal ID'] / merged['Total'] * 100).round(1)
        merged = merged.sort_values('No Deal ID', ascending=False)

        worst5 = merged.head(5)
        cols = st.columns(min(len(worst5), 5))
        for i, (_, row) in enumerate(worst5.iterrows()):
            pct = row['% Untracked']
            color = '#E74C3C' if pct > 50 else ('#FFC000' if pct > 20 else '#70AD47')
            with cols[i]:
                st.markdown(f"""
                <div style="background:rgba(255,255,255,0.04);border-radius:12px;padding:1rem;text-align:center;border:1px solid rgba(255,255,255,0.08);">
                    <div style="font-size:1.8rem;font-weight:800;color:{color};">H{int(row['Hanger#'])}</div>
                    <div style="font-size:1.3rem;font-weight:700;color:white;">{int(row['No Deal ID'])}</div>
                    <div style="color:#aaaacc;font-size:0.8rem;">untracked / {int(row['Total'])} total</div>
                    <div style="color:{color};font-size:0.85rem;font-weight:600;">{pct}% untracked</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)
        fig_nd = go.Figure(go.Bar(
            x=merged['Hanger#'].astype(int).astype(str),
            y=merged['No Deal ID'],
            marker=dict(
                color=merged['% Untracked'],
                colorscale=[[0,'#70AD47'],[0.4,'#FFC000'],[1,'#E74C3C']],
                showscale=True,
                colorbar=dict(title='% Untracked', tickfont=dict(color='white'), title_font=dict(color='white')),
            ),
            text=merged['No Deal ID'],
            textposition='outside',
            textfont=dict(color='white'),
            hovertemplate='Hanger %{x}<br>%{y} untracked devices<br>%{customdata}% untracked<extra></extra>',
            customdata=merged['% Untracked'],
        ))
        fig_nd.update_layout(
            height=350, margin=dict(l=10,r=10,t=20,b=10),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(title='Hanger', color='white', showgrid=False),
            yaxis=dict(title='Untracked Count', color='white', showgrid=False),
            font=dict(color='white'),
        )
        st.plotly_chart(fig_nd, use_container_width=True)

        with st.expander("📋 Full table — all hangers"):
            st.dataframe(merged.rename(columns={'Hanger#': 'Hanger'}).astype({'Hanger': int}),
                         use_container_width=True, hide_index=True)
    else:
        st.info("Generate the Master Template first to see this report.")

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 3 — GRADE DISTRIBUTION PER HANGER
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="diag-section">', unsafe_allow_html=True)
    st.markdown('<div class="diag-title">🎯 Grade Distribution per Hanger</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-sub">Quality breakdown per hanger — know which hangers hold premium vs low-grade stock at a glance</div>', unsafe_allow_html=True)

    if master_df is not None and hanger_raw is not None:
        hm = master_df[(master_df['Bin'] == 'Hanger') & (master_df['Grade'] != '')].copy()
        hm['Hanger#'] = hm['Location'].str.extract(r'^(\d+)').astype(float)
        grade_pivot = hm.groupby(['Hanger#', 'Grade']).size().reset_index(name='Count')

        all_grades_h = sorted(grade_pivot['Grade'].unique())
        PALETTE = ['#11998e','#70AD47','#4472C4','#FFC000','#E74C3C',
                   '#8E44AD','#EC6EAD','#3494E6','#f5576c','#38ef7d',
                   '#FF6B35','#A8E6CF','#FFD93D','#6BCB77','#f093fb']
        g_color = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(all_grades_h)}

        hangers_sorted = sorted(grade_pivot['Hanger#'].dropna().unique())
        traces = []
        for grade in all_grades_h:
            sub = grade_pivot[grade_pivot['Grade'] == grade]
            traces.append(go.Bar(
                name=grade,
                x=sub['Hanger#'].astype(int).astype(str),
                y=sub['Count'],
                marker_color=g_color[grade],
                hovertemplate=f'<b>{grade}</b><br>Hanger %{{x}}<br>%{{y}} devices<extra></extra>',
            ))

        fig_gh = go.Figure(traces)
        fig_gh.update_layout(
            barmode='stack',
            height=400, margin=dict(l=10,r=10,t=20,b=10),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(title='Hanger', color='white', showgrid=False, tickangle=-45),
            yaxis=dict(title='Device Count', color='white', showgrid=False),
            legend=dict(font=dict(color='white', size=11), orientation='h', y=-0.2),
            font=dict(color='white'),
        )
        st.plotly_chart(fig_gh, use_container_width=True)
    else:
        st.info("Generate the Master Template first to see this report.")

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 4 — TOTES ANALYTICS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="diag-section">', unsafe_allow_html=True)
    st.markdown('<div class="diag-title">🗂️ Totes Analytics</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-sub">Visibility into totes storage — currently a blind spot in the main dashboard</div>', unsafe_allow_html=True)

    if totes_raw is not None:
        real_totes = totes_raw[totes_raw['IMEI'].apply(_is_real_device)]
        empty_totes = totes_raw[~totes_raw['IMEI'].apply(_is_real_device)]

        t1, t2, t3, t4 = st.columns(4)
        for col, val, lbl, color in [
            (t1, len(real_totes),  'Devices in Totes', '#4472C4'),
            (t2, len(empty_totes), 'Empty Tote Slots',  '#FFC000'),
            (t3, int(real_totes['LOA'].nunique()), 'Unique Tote IDs', '#70AD47'),
            (t4, int((real_totes['Deal ID'].astype(str).str.contains('AE', na=False) == False).sum()), 'No Deal ID', '#E74C3C'),
        ]:
            with col:
                st.markdown(f"""
                <div style="background:rgba(255,255,255,0.04);border-radius:12px;padding:1rem;text-align:center;border:1px solid rgba(255,255,255,0.08);">
                    <div style="font-size:2rem;font-weight:800;color:{color};">{val:,}</div>
                    <div style="color:#aaaacc;font-size:0.82rem;">{lbl}</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)

        if master_df is not None:
            totes_master = master_df[master_df['Bin'] == 'Totes']
            if len(totes_master) > 0:
                tc1, tc2 = st.columns(2)

                with tc1:
                    # devices per tote location
                    tote_loc = totes_master.groupby('Location').size().reset_index(name='Count').sort_values('Count', ascending=False)
                    fig_tl = go.Figure(go.Bar(
                        x=tote_loc['Location'], y=tote_loc['Count'],
                        marker=dict(color=tote_loc['Count'],
                                    colorscale=[[0,'#1a1a2e'],[0.5,'#4472C4'],[1,'#70AD47']]),
                        text=tote_loc['Count'], textposition='outside',
                        textfont=dict(color='white'),
                        hovertemplate='Tote %{x}<br>%{y} devices<extra></extra>',
                    ))
                    fig_tl.update_layout(
                        title=dict(text='Devices per Tote', font=dict(color='white', size=13), x=0.5),
                        height=300, margin=dict(l=10,r=10,t=40,b=10),
                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                        xaxis=dict(color='white', showgrid=False),
                        yaxis=dict(color='white', showgrid=False),
                    )
                    st.plotly_chart(fig_tl, use_container_width=True)

                with tc2:
                    # grade breakdown totes
                    tg = totes_master[totes_master['Grade'] != '']['Grade'].value_counts().reset_index()
                    tg.columns = ['Grade', 'Count']
                    PALETTE = ['#11998e','#70AD47','#4472C4','#FFC000','#E74C3C','#8E44AD','#EC6EAD','#f5576c']
                    tg_colors = [PALETTE[i % len(PALETTE)] for i in range(len(tg))]
                    fig_tg = go.Figure(go.Pie(
                        labels=tg['Grade'], values=tg['Count'], hole=0.5,
                        marker=dict(colors=tg_colors, line=dict(color='rgba(0,0,0,0)', width=0)),
                        textfont=dict(color='white', size=11),
                        hovertemplate='%{label}<br>%{value} devices (%{percent})<extra></extra>',
                    ))
                    fig_tg.update_layout(
                        title=dict(text='Grade Breakdown (Totes)', font=dict(color='white', size=13), x=0.5),
                        height=300, margin=dict(l=10,r=10,t=40,b=10),
                        paper_bgcolor='rgba(0,0,0,0)',
                        legend=dict(font=dict(color='white', size=10)),
                    )
                    st.plotly_chart(fig_tg, use_container_width=True)
    else:
        st.info("Upload a Totes file to see totes analytics.")

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 5 — MISMATCHED / ERROR DEAL IDs
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="diag-section">', unsafe_allow_html=True)
    st.markdown('<div class="diag-title">⚠️ Mismatched & Error Deal IDs</div>', unsafe_allow_html=True)
    st.markdown('<div class="diag-sub">Devices with real IMEIs but problematic Deal IDs — formula errors, non-AE codes, #N/A values that slip through unnoticed</div>', unsafe_allow_html=True)

    error_rows = []
    for src, raw in [('Hanger', hanger_raw), ('Totes', totes_raw)]:
        if raw is None: continue
        for _, row in raw.iterrows():
            if not _is_real_device(row.get('IMEI', '')): continue
            imei      = _clean_imei(row['IMEI'])
            deal_id   = str(row.get('Deal ID', '')).strip()
            backend   = str(row.get('BACKEND', '')).strip()
            loc       = (str(int(row['Hangar'])) if src == 'Hanger' and pd.notna(row.get('Hangar')) else '') + str(row.get('LOA', '')).strip()
            issues    = []
            # Rule 1: neither Deal ID nor BACKEND has AE
            has_ae_deal = 'AE' in deal_id.upper()
            has_ae_back = 'AE' in backend.upper()
            if not has_ae_deal and not has_ae_back and deal_id.upper() not in ('', 'NAN'):
                issues.append('Non-AE Deal ID')
            # Rule 2: #N/A or formula error
            if '#N/A' in deal_id or '#N/A' in backend or 'REF' in backend.upper():
                issues.append('Formula Error (#N/A / #REF)')
            # Rule 3: BACKEND blank but Deal ID also blank
            if deal_id in ('', 'nan') and backend in ('', 'nan', '0'):
                issues.append('Both Deal ID & BACKEND missing')
            if issues:
                error_rows.append({'Source': src, 'Location': loc, 'IMEI': imei,
                                   'Deal ID': deal_id, 'BACKEND': backend,
                                   'Issue': ' | '.join(issues)})

    if error_rows:
        err_df = pd.DataFrame(error_rows)
        st.markdown(f'<span class="badge-red">⚠️ {len(err_df)} rows flagged</span>&nbsp;&nbsp;', unsafe_allow_html=True)
        issue_counts = err_df['Issue'].value_counts().reset_index()
        issue_counts.columns = ['Issue Type', 'Count']
        ic1, ic2 = st.columns([1, 2])
        with ic1:
            for _, ir in issue_counts.iterrows():
                st.markdown(f'<div style="background:rgba(231,76,60,0.1);border-left:3px solid #E74C3C;padding:0.5rem 1rem;border-radius:4px;margin:0.4rem 0;">'
                            f'<b style="color:#E74C3C;">{int(ir["Count"])}</b> &nbsp;<span style="color:white;">{ir["Issue Type"]}</span></div>',
                            unsafe_allow_html=True)
        with ic2:
            st.dataframe(err_df, use_container_width=True, hide_index=True)
    else:
        st.markdown('<span class="badge-green">✅ No mismatched or error Deal IDs found</span>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# DAILY DELTA REPORT — standalone function
# ══════════════════════════════════════════════════════════════════════════════
def show_daily_delta(hanger_file, totes_file):
    hanger_raw = None
    totes_raw  = None

    if hanger_file:
        try:
            hanger_file.seek(0)
            hanger_raw = pd.read_excel(hanger_file, sheet_name='Sheet1', engine='openpyxl')
            hanger_raw = hanger_raw[hanger_raw['Room'].notna()].reset_index(drop=True)
        except Exception:
            hanger_raw = None

    if totes_file:
        try:
            totes_file.seek(0)
            totes_raw = pd.read_excel(totes_file, sheet_name='Sheet1', engine='openpyxl')
            totes_raw = totes_raw[totes_raw['Room'].notna()].reset_index(drop=True)
        except Exception:
            totes_raw = None

    import plotly.graph_objects as go

    st.markdown("""
    <style>
    .delta-section{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);
                   border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;}
    .delta-title{font-size:1.25rem;font-weight:700;color:#ffffff;margin-bottom:0.3rem;}
    .delta-sub{font-size:0.85rem;color:#aaaacc;margin-bottom:1rem;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="delta-section">', unsafe_allow_html=True)
    st.markdown('<div class="delta-title">📅 Daily Comparison — What Changed?</div>', unsafe_allow_html=True)
    st.markdown('<div class="delta-sub">Upload yesterday\'s stocktake files to automatically see what moved, what\'s new, and what disappeared — eliminates manual daily reconciliation</div>', unsafe_allow_html=True)

    dc1, dc2 = st.columns(2)
    with dc1:
        prev_hanger = st.file_uploader("📁 Yesterday's Hanger file", type=['xlsx','xlsm','xls'], key='prev_hanger_delta')
    with dc2:
        prev_totes  = st.file_uploader("🗂️ Yesterday's Totes file",  type=['xlsx','xlsm','xls'], key='prev_totes_delta')

    if prev_hanger or prev_totes:
        _delta_result = {}
        with st.spinner("Processing and comparing files — please wait..."):
            # build today's IMEI→location map
            today_map = {}
            if hanger_raw is not None:
                for _, row in hanger_raw.iterrows():
                    if not _is_real_device(row.get('IMEI','')): continue
                    h_val = row.get('Hangar', '')
                    try:
                        h_str = str(int(float(str(h_val).strip()))) if pd.notna(h_val) else ''
                    except (ValueError, TypeError):
                        continue
                    imei = _clean_imei(row['IMEI'])
                    loc  = h_str + str(row.get('LOA','')).strip()
                    today_map[imei] = ('Hanger', loc)
            if totes_raw is not None:
                for _, row in totes_raw.iterrows():
                    if not _is_real_device(row.get('IMEI','')): continue
                    imei = _clean_imei(row['IMEI'])
                    loc  = str(row.get('LOA','')).strip()
                    today_map[imei] = ('Totes', loc)

            # build yesterday's IMEI→location map
            yest_map = {}
            for src, pf in [('Hanger', prev_hanger), ('Totes', prev_totes)]:
                if pf is None: continue
                pf.seek(0)
                try:
                    pdf = pd.read_excel(pf, sheet_name='Sheet1', engine='openpyxl')
                    pdf = pdf[pdf['Room'].notna()].reset_index(drop=True)
                    for _, row in pdf.iterrows():
                        if not _is_real_device(row.get('IMEI','')): continue
                        imei = _clean_imei(row['IMEI'])
                        if src == 'Hanger':
                            loc = (str(int(row['Hangar'])) if pd.notna(row.get('Hangar')) else '') + str(row.get('LOA','')).strip()
                        else:
                            loc = str(row.get('LOA','')).strip()
                        yest_map[imei] = (src, loc)
                except Exception as e:
                    st.error(f"Could not read previous {src} file: {e}")

            new_devices  = {k: today_map[k] for k in today_map if k not in yest_map}
            removed_devs = {k: yest_map[k]  for k in yest_map  if k not in today_map}
            moved_devs   = {k: (yest_map[k], today_map[k]) for k in today_map
                            if k in yest_map and yest_map[k] != today_map[k]}

        d1, d2, d3 = st.columns(3)
        for col, count, lbl, color in [
            (d1, len(new_devices),  'New Devices Added',    '#70AD47'),
            (d2, len(removed_devs), 'Devices Removed',      '#E74C3C'),
            (d3, len(moved_devs),   'Devices Moved/Shifted','#FFC000'),
        ]:
            with col:
                st.markdown(f"""
                <div style="background:rgba(255,255,255,0.04);border-radius:12px;padding:1.2rem;text-align:center;border:1px solid rgba(255,255,255,0.08);">
                    <div style="font-size:2.5rem;font-weight:800;color:{color};">{count}</div>
                    <div style="color:#aaaacc;font-size:0.85rem;">{lbl}</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)

        # ── VISUALS ──────────────────────────────────────────────────────────
        import re as _re

        def _hanger_num(loc):
            m = _re.match(r'^(\d+)', str(loc))
            return m.group(1) if m else None

        # Per-hanger counts today vs yesterday
        today_h_cnt, yest_h_cnt = {}, {}
        today_totes_cnt = yest_totes_cnt = 0
        for imei, (typ, loc) in today_map.items():
            if typ == 'Hanger':
                h = _hanger_num(loc)
                if h: today_h_cnt[h] = today_h_cnt.get(h, 0) + 1
            else:
                today_totes_cnt += 1
        for imei, (typ, loc) in yest_map.items():
            if typ == 'Hanger':
                h = _hanger_num(loc)
                if h: yest_h_cnt[h] = yest_h_cnt.get(h, 0) + 1
            else:
                yest_totes_cnt += 1

        all_hangers = sorted(set(list(today_h_cnt.keys()) + list(yest_h_cnt.keys())), key=lambda x: int(x))
        yest_vals   = [yest_h_cnt.get(h, 0) for h in all_hangers]
        today_vals  = [today_h_cnt.get(h, 0) for h in all_hangers]
        diff_vals   = [t - y for t, y in zip(today_vals, yest_vals)]
        bar_colors  = ['#70AD47' if d >= 0 else '#E74C3C' for d in diff_vals]

        vcol1, vcol2 = st.columns([2, 1])

        with vcol1:
            if all_hangers:
                fig_cmp = go.Figure()
                fig_cmp.add_trace(go.Bar(
                    name='Yesterday', x=all_hangers, y=yest_vals,
                    marker_color='rgba(100,100,180,0.55)',
                    hovertemplate='Hanger %{x}<br>Yesterday: %{y}<extra></extra>',
                ))
                fig_cmp.add_trace(go.Bar(
                    name='Today', x=all_hangers, y=today_vals,
                    marker_color='rgba(68,114,196,0.9)',
                    hovertemplate='Hanger %{x}<br>Today: %{y}<extra></extra>',
                ))
                fig_cmp.add_trace(go.Scatter(
                    name='Net Change', x=all_hangers, y=diff_vals,
                    mode='lines+markers',
                    line=dict(color='#FFC000', width=2, dash='dot'),
                    marker=dict(color=bar_colors, size=8, line=dict(color='white', width=1)),
                    hovertemplate='Hanger %{x}<br>Net: %{y:+d}<extra></extra>',
                    yaxis='y2',
                ))
                fig_cmp.update_layout(
                    title=dict(text='📊 Hanger Device Count — Yesterday vs Today',
                               font=dict(color='white', size=14), x=0.02),
                    barmode='group', height=340,
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=10, r=10, t=45, b=30),
                    legend=dict(font=dict(color='white', size=11),
                                bgcolor='rgba(255,255,255,0.05)', bordercolor='rgba(255,255,255,0.1)', borderwidth=1),
                    xaxis=dict(title='Hanger', color='white', showgrid=False, tickfont=dict(size=10)),
                    yaxis=dict(title='Devices', color='white', showgrid=True,
                               gridcolor='rgba(255,255,255,0.06)'),
                    yaxis2=dict(title='Net Δ', color='#FFC000', overlaying='y', side='right',
                                showgrid=False, zeroline=True, zerolinecolor='rgba(255,255,255,0.15)'),
                    font=dict(color='white'),
                )
                st.plotly_chart(fig_cmp, use_container_width=True)

        with vcol2:
            total_changes = len(new_devices) + len(removed_devs) + len(moved_devs)
            if total_changes > 0:
                chg_labels = ['New Devices', 'Removed', 'Moved/Shifted']
                chg_vals   = [len(new_devices), len(removed_devs), len(moved_devs)]
                chg_colors = ['#70AD47', '#E74C3C', '#FFC000']
                chg_vals_nz   = [v for v in chg_vals if v > 0]
                chg_labels_nz = [l for l, v in zip(chg_labels, chg_vals) if v > 0]
                chg_colors_nz = [c for c, v in zip(chg_colors, chg_vals) if v > 0]

                fig_donut = go.Figure(go.Pie(
                    labels=chg_labels_nz, values=chg_vals_nz,
                    hole=0.62,
                    marker=dict(colors=chg_colors_nz,
                                line=dict(color='rgba(0,0,0,0.4)', width=2)),
                    textfont=dict(color='white', size=11),
                    hovertemplate='%{label}<br>%{value} devices (%{percent})<extra></extra>',
                    direction='clockwise', sort=False,
                ))
                fig_donut.add_annotation(
                    text=f"<b>{total_changes}</b><br><span style='font-size:11px'>changes</span>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(color='white', size=18),
                    align='center',
                )
                fig_donut.update_layout(
                    title=dict(text='Change Breakdown', font=dict(color='white', size=13), x=0.5),
                    height=340, margin=dict(l=10, r=10, t=45, b=10),
                    paper_bgcolor='rgba(0,0,0,0)',
                    legend=dict(font=dict(color='white', size=11), orientation='v',
                                bgcolor='rgba(255,255,255,0.04)'),
                    showlegend=True,
                )
                st.plotly_chart(fig_donut, use_container_width=True)
            else:
                st.markdown("""
                <div style="height:340px;display:flex;align-items:center;justify-content:center;
                            background:rgba(112,173,71,0.08);border:1px solid rgba(112,173,71,0.2);
                            border-radius:14px;text-align:center;padding:2rem;">
                    <div>
                        <div style="font-size:3rem;">✅</div>
                        <div style="color:#70AD47;font-weight:700;font-size:1.1rem;margin-top:0.5rem;">No Changes</div>
                        <div style="color:#aaaacc;font-size:0.85rem;margin-top:0.3rem;">Inventory identical to yesterday</div>
                    </div>
                </div>""", unsafe_allow_html=True)

        # ── Movement summary strip ────────────────────────────────────────────
        if moved_devs:
            move_types = {}
            for imei, (yest, today) in moved_devs.items():
                key = f"{yest[0]} → {today[0]}"
                move_types[key] = move_types.get(key, 0) + 1
            mcols = st.columns(len(move_types))
            icons = {'Hanger → Hanger': '🔀', 'Hanger → Totes': '📦', 'Totes → Hanger': '🏷️', 'Totes → Totes': '🔄'}
            colors_mv = {'Hanger → Hanger': '#4472C4', 'Hanger → Totes': '#FFC000', 'Totes → Hanger': '#70AD47', 'Totes → Totes': '#8E44AD'}
            for col, (mv_type, mv_cnt) in zip(mcols, move_types.items()):
                ic   = icons.get(mv_type, '↔️')
                clr  = colors_mv.get(mv_type, '#aaaacc')
                with col:
                    st.markdown(f"""
                    <div style="background:{clr}15;border:1px solid {clr}40;border-radius:12px;
                                padding:0.9rem;text-align:center;margin-bottom:1rem;">
                        <div style="font-size:1.8rem;">{ic}</div>
                        <div style="font-size:1.5rem;font-weight:800;color:{clr};">{mv_cnt}</div>
                        <div style="font-size:0.78rem;color:#aaaacc;">{mv_type}</div>
                    </div>""", unsafe_allow_html=True)

        st.markdown('<hr style="border-color:rgba(255,255,255,0.08);margin:0.5rem 0 1.2rem;">', unsafe_allow_html=True)

        def _delta_excel(df):
            import io, xlsxwriter as xw

            df = df.copy()
            if 'IMEI' in df.columns:
                df['IMEI'] = df['IMEI'].apply(
                    lambda v: str(v).strip().split('.')[0] if pd.notna(v) else ''
                )

            out = io.BytesIO()
            wb = xw.Workbook(out, {'in_memory': True})
            ws = wb.add_worksheet('Delta')

            hdr_fmt  = wb.add_format({'bold': True, 'font_color': 'white',
                                      'bg_color': '4472C4', 'align': 'center',
                                      'valign': 'vcenter', 'border': 1})
            str_fmt  = wb.add_format({'num_format': '@', 'border': 1})
            cell_fmt = wb.add_format({'border': 1})

            cols = list(df.columns)
            imei_col_idx = cols.index('IMEI') if 'IMEI' in cols else None

            ws.set_row(0, 25)
            ws.freeze_panes(1, 0)
            for c_idx, col_name in enumerate(cols):
                ws.set_column(c_idx, c_idx, 22)
                ws.write(0, c_idx, col_name, hdr_fmt)

            for r_idx, row in enumerate(df.itertuples(index=False), 1):
                for c_idx, val in enumerate(row):
                    if c_idx == imei_col_idx:
                        ws.write_string(r_idx, c_idx, str(val), str_fmt)
                    else:
                        ws.write(r_idx, c_idx, val, cell_fmt)

            wb.close()
            out.seek(0)
            return out

        tab_new, tab_rem, tab_moved = st.tabs([f"✅ New ({len(new_devices)})", f"❌ Removed ({len(removed_devs)})", f"🔄 Moved ({len(moved_devs)})"])

        with tab_new:
            if new_devices:
                nd = pd.DataFrame([{'IMEI': k, 'Storage Type': v[0], 'Location': v[1]} for k, v in new_devices.items()])
                nd['IMEI'] = nd['IMEI'].astype(str)
                st.caption("⚠️ Use the download button below — do NOT use the icon inside the table (that exports CSV which breaks IMEI format)")
                st.dataframe(nd, use_container_width=True, hide_index=True)
                st.download_button("⬇️ Download New Devices (Excel — IMEI formatted correctly)",
                    data=_delta_excel(nd),
                    file_name="Delta_New_Devices.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='dl_new', type='primary', use_container_width=True)
            else:
                st.success("No new devices since yesterday.")

        with tab_rem:
            if removed_devs:
                rd = pd.DataFrame([{'IMEI': k, 'Last Seen In': v[0], 'Last Location': v[1]} for k, v in removed_devs.items()])
                rd['IMEI'] = rd['IMEI'].astype(str)
                st.caption("⚠️ Use the download button below — do NOT use the icon inside the table (that exports CSV which breaks IMEI format)")
                st.dataframe(rd, use_container_width=True, hide_index=True)
                st.download_button("⬇️ Download Removed Devices (Excel — IMEI formatted correctly)",
                    data=_delta_excel(rd),
                    file_name="Delta_Removed_Devices.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='dl_rem', type='primary', use_container_width=True)
            else:
                st.success("No devices removed since yesterday.")

        with tab_moved:
            if moved_devs:
                md_rows = []
                for imei, (yest, today) in moved_devs.items():
                    md_rows.append({'IMEI': imei,
                                    'From Type': yest[0], 'From Location': yest[1],
                                    'To Type': today[0],  'To Location': today[1]})
                mdf = pd.DataFrame(md_rows)
                mdf['IMEI'] = mdf['IMEI'].astype(str)
                st.caption("⚠️ Use the download button below — do NOT use the icon inside the table (that exports CSV which breaks IMEI format)")
                st.dataframe(mdf, use_container_width=True, hide_index=True)
                st.download_button("⬇️ Download Moved Devices (Excel — IMEI formatted correctly)",
                    data=_delta_excel(mdf),
                    file_name="Delta_Moved_Devices.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='dl_moved', type='primary', use_container_width=True)
            else:
                st.success("No devices changed location since yesterday.")
    else:
        st.markdown("""
        <div style="background:rgba(255,255,255,0.03);border:1px dashed rgba(255,255,255,0.15);
                    border-radius:12px;padding:2rem;text-align:center;color:#aaaacc;">
            Upload yesterday's stocktake files above to see the delta report
        </div>
        """, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# STOCK INTELLIGENCE — Distribution Channels + Low Value
# ══════════════════════════════════════════════════════════════════════════════
def show_stock_intelligence(stack_file):
    import plotly.graph_objects as go
    import plotly.express as px

    st.markdown(
        '<div class="section-banner">'
        '<div class="banner-title">📦 Stock Intelligence</div>'
        '<div class="banner-sub">Distribution channels & low value inventory — TV and Home Appliances excluded (stored externally)</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    with st.spinner("Reading Stack Bulk..."):
        stack_df, warn = read_stack_bulk(stack_file)
    if stack_df is None:
        st.error("Could not read Stack Bulk file.")
        return
    if warn:
        st.info(warn)

    # ── Normalise QR Code Description ────────────────────────────────────────
    stack_df = stack_df.copy()
    stack_df['_qr'] = stack_df['QR Code Description'].astype(str).str.upper().str.strip()
    stack_df['_cat'] = stack_df['Category'].astype(str).str.strip()

    # Exclude TV and Home Appliances
    external_mask = stack_df['_cat'].str.lower().isin(['tv', 'home appliances'])
    external_count = external_mask.sum()
    stack_df = stack_df[~external_mask].copy()

    st.caption(f"ℹ️ {external_count} TV / Home Appliance items excluded — stored outside warehouse")

    # ── Define channel groups ─────────────────────────────────────────────────
    DT_CHANNELS = {
        'DT1': {'label': 'DT1 — Dealer Team',          'color': '#4472C4'},
        'DT2': {'label': 'DT2 — Abu Dhabi Dealers',    'color': '#ED7D31'},
        'DT3': {'label': 'DT3 — E-Commerce',           'color': '#70AD47'},
        'DT4': {'label': 'DT4 — Refurbishment',        'color': '#FFC000'},
        'DT5': {'label': 'DT5 — Branded Delivery',     'color': '#FF4B4B'},
    }
    LOW_VALUE = {
        'LVIN1': {'label': 'LVIN1 — Low Value 1', 'color': '#9B59B6'},
        'LVIN2': {'label': 'LVIN2 — Low Value 2', 'color': '#E91E63'},
        'R151':  {'label': 'R151 — Low Value Batch', 'color': '#00BCD4'},
    }

    def get_subset(code):
        return stack_df[stack_df['_qr'] == code].copy()

    # ══ SECTION 1: DISTRIBUTION CHANNELS ═════════════════════════════════════
    st.markdown("### 🚚 Distribution Channels")

    # Summary scorecards
    dt_cols = st.columns(5)
    for i, (code, meta) in enumerate(DT_CHANNELS.items()):
        sub = get_subset(code)
        top_brand = sub['Brand'].value_counts().index[0] if len(sub) > 0 else '—'
        top_grade = sub['Sell Grade'].astype(str).str.strip().str.rstrip('.').value_counts().index[0] if len(sub) > 0 else '—'
        with dt_cols[i]:
            st.markdown(f"""
            <div style='background:{meta["color"]}22;border:1px solid {meta["color"]};
                        border-radius:12px;padding:1rem;text-align:center;'>
              <div style='font-size:1.8rem;font-weight:700;color:{meta["color"]};'>{len(sub)}</div>
              <div style='font-size:0.75rem;color:white;font-weight:600;margin:0.2rem 0;'>{meta["label"]}</div>
              <div style='font-size:0.7rem;color:#aaa;'>Top: {top_brand}</div>
              <div style='font-size:0.7rem;color:#aaa;'>Grade: {top_grade}</div>
            </div>
            """, unsafe_allow_html=True)

    grade_colors = {
        'Grade A-Plus':'#00C851','Grade A':'#4CAF50','Grade B':'#8BC34A',
        'Grade C':'#FFC107','Grade D1':'#FF9800','Grade D2':'#FF5722',
        'Grade F':'#F44336','Average':'#9E9E9E','Flawless':'#00BCD4'
    }

    st.markdown("<br>", unsafe_allow_html=True)

    def hex_to_rgba(hex_col, alpha):
        h = hex_col.lstrip('#')
        r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
        return f'rgba({r},{g},{b},{alpha})'

    tab1, tab2, tab3 = st.tabs(["🗺️ Treemap — Channel · Brand · Grade", "🌊 Sankey — Channel → Brand → Country", "⬇️ Download by Channel"])

    with tab1:
        # Treemap: Channel → Brand → Grade, sized by device count
        tm_ids, tm_labels, tm_parents, tm_values, tm_colors = [], [], [], [], []

        all_dt = pd.concat([get_subset(c) for c in DT_CHANNELS], ignore_index=True)
        tm_ids.append('ALL'); tm_labels.append(f'All Channels<br>{len(all_dt)} devices')
        tm_parents.append(''); tm_values.append(len(all_dt)); tm_colors.append('#1a1a2e')

        for code, meta in DT_CHANNELS.items():
            sub = get_subset(code)
            if len(sub) == 0:
                continue
            tm_ids.append(code)
            tm_labels.append(f"{meta['label']}<br>{len(sub)} devices")
            tm_parents.append('ALL'); tm_values.append(len(sub)); tm_colors.append(meta['color'])

            for brand, b_cnt in sub['Brand'].value_counts().head(8).items():
                b_id = f"{code}_{brand}"
                tm_ids.append(b_id); tm_labels.append(f"{brand}<br>{int(b_cnt)}")
                tm_parents.append(code); tm_values.append(int(b_cnt))
                tm_colors.append(hex_to_rgba(meta['color'], 0.75))

                brand_sub = sub[sub['Brand'] == brand]
                for grade, g_cnt in brand_sub['Sell Grade'].astype(str).str.strip().str.rstrip('.').value_counts().items():
                    g_id = f"{code}_{brand}_{grade}"
                    tm_ids.append(g_id); tm_labels.append(f"{grade}<br>{int(g_cnt)}")
                    tm_parents.append(b_id); tm_values.append(int(g_cnt))
                    tm_colors.append(grade_colors.get(grade, '#888888'))

        fig_tm = go.Figure(go.Treemap(
            ids=tm_ids, labels=tm_labels, parents=tm_parents,
            values=tm_values, marker=dict(colors=tm_colors),
            branchvalues='total',
            hovertemplate='<b>%{label}</b><br>Devices: %{value}<br>Share of parent: %{percentParent:.1%}<extra></extra>',
            textfont=dict(size=12, color='white'),
            pathbar=dict(visible=True, thickness=20),
        ))
        fig_tm.update_layout(
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(t=30, b=10, l=10, r=10), height=580,
        )
        st.plotly_chart(fig_tm, use_container_width=True)
        st.caption("Click any block to drill in. Use the breadcrumb bar at top to navigate back.")

    with tab2:
        # Sankey: Channel → Brand → Country
        all_nodes = []
        def node_idx(name):
            if name not in all_nodes:
                all_nodes.append(name)
            return all_nodes.index(name)

        sk_src, sk_tgt, sk_val, sk_colors_link = [], [], [], []

        for code, meta in DT_CHANNELS.items():
            sub = get_subset(code)
            if len(sub) == 0:
                continue
            c_idx = node_idx(code)

            for brand, b_grp in sub.groupby('Brand'):
                brand = str(brand)
                b_idx = node_idx(brand)
                sk_src.append(c_idx); sk_tgt.append(b_idx); sk_val.append(len(b_grp))
                sk_colors_link.append(hex_to_rgba(meta['color'], 0.55))

                for country, cnt in b_grp['Storage Member Country'].value_counts().items():
                    co_idx = node_idx(str(country))
                    sk_src.append(b_idx); sk_tgt.append(co_idx); sk_val.append(int(cnt))
                    sk_colors_link.append(hex_to_rgba(meta['color'], 0.28))

        node_colors = []
        for n in all_nodes:
            if n in DT_CHANNELS:
                node_colors.append(DT_CHANNELS[n]['color'])
            else:
                node_colors.append('#888888')

        fig_sk = go.Figure(go.Sankey(
            arrangement='snap',
            node=dict(
                pad=18, thickness=22,
                label=all_nodes,
                color=node_colors,
                hovertemplate='<b>%{label}</b><br>%{value} devices<extra></extra>',
            ),
            link=dict(
                source=sk_src, target=sk_tgt, value=sk_val,
                color=sk_colors_link,
                hovertemplate='%{source.label} → %{target.label}<br>%{value} devices<extra></extra>',
            )
        ))
        fig_sk.update_layout(
            paper_bgcolor='rgba(0,0,0,0)', font_color='white',
            font_size=11, margin=dict(t=10, b=10, l=10, r=10), height=560,
        )
        st.plotly_chart(fig_sk, use_container_width=True)
        st.caption("Flow width = device volume. Left = channels, middle = brands, right = destination countries.")

    with tab3:
        st.markdown("**Download full data for each channel**")
        dl_cols = st.columns(5)
        for i, (code, meta) in enumerate(DT_CHANNELS.items()):
            sub = get_subset(code).drop(columns=['_qr', '_cat'], errors='ignore')
            cache_key = f'_si_dt_{code}'
            if cache_key not in st.session_state:
                buf = BytesIO()
                import xlsxwriter as xw
                with xw.Workbook(buf, {'in_memory': True}) as wb:
                    ws = wb.add_worksheet(code)
                    hdr_fmt = wb.add_format({'bold': True, 'bg_color': '#1F4E79', 'font_color': 'white', 'border': 1})
                    txt_fmt = wb.add_format({'num_format': '@'})
                    for ci, col_name in enumerate(sub.columns):
                        ws.write(0, ci, col_name, hdr_fmt)
                        ws.set_column(ci, ci, 20)
                    for ri, row_data in enumerate(sub.itertuples(index=False), start=1):
                        for ci, val in enumerate(row_data):
                            if sub.columns[ci] == 'IMEI Number':
                                ws.write_string(ri, ci, str(val).strip().split('.')[0] if pd.notna(val) else '', txt_fmt)
                            else:
                                ws.write(ri, ci, val if pd.notna(val) else '')
                buf.seek(0)
                st.session_state[cache_key] = buf.read()
            with dl_cols[i]:
                st.download_button(
                    f"⬇️ {code}",
                    data=st.session_state[cache_key],
                    file_name=f"Stock_{code}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                    key=f'dl_dt_{code}'
                )
                st.caption(f"{len(sub):,} rows")

    st.markdown("---")

    # ══ SECTION 2: LOW VALUE INVENTORY ═══════════════════════════════════════
    st.markdown("### 🏷️ Low Value Inventory")

    lv_cols = st.columns(3)
    for i, (code, meta) in enumerate(LOW_VALUE.items()):
        sub = get_subset(code)
        top_brand = sub['Brand'].value_counts().index[0] if len(sub) > 0 else '—'
        top_cat = sub['Category'].value_counts().index[0] if len(sub) > 0 else '—'
        with lv_cols[i]:
            st.markdown(f"""
            <div style='background:{meta["color"]}22;border:1px solid {meta["color"]};
                        border-radius:12px;padding:1rem;text-align:center;'>
              <div style='font-size:1.8rem;font-weight:700;color:{meta["color"]};'>{len(sub)}</div>
              <div style='font-size:0.75rem;color:white;font-weight:600;margin:0.2rem 0;'>{meta["label"]}</div>
              <div style='font-size:0.7rem;color:#aaa;'>Brand: {top_brand}</div>
              <div style='font-size:0.7rem;color:#aaa;'>Type: {top_cat}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    lv_tab1, lv_tab2 = st.tabs(["🗺️ Treemap — Group · Brand · Grade", "⬇️ Download Low Value"])

    with lv_tab1:
        lv_ids, lv_labels, lv_parents, lv_values, lv_clrs = [], [], [], [], []

        all_lv = pd.concat([get_subset(c) for c in LOW_VALUE], ignore_index=True)
        lv_ids.append('LV_ALL'); lv_labels.append(f'Low Value<br>{len(all_lv)} devices')
        lv_parents.append(''); lv_values.append(len(all_lv)); lv_clrs.append('#1a1a2e')

        for code, meta in LOW_VALUE.items():
            sub = get_subset(code)
            if len(sub) == 0:
                continue
            lv_ids.append(code); lv_labels.append(f"{meta['label']}<br>{len(sub)}")
            lv_parents.append('LV_ALL'); lv_values.append(len(sub)); lv_clrs.append(meta['color'])

            for brand, b_cnt in sub['Brand'].value_counts().items():
                b_id = f"lv_{code}_{brand}"
                lv_ids.append(b_id); lv_labels.append(f"{brand}<br>{int(b_cnt)}")
                lv_parents.append(code); lv_values.append(int(b_cnt))
                lv_clrs.append(hex_to_rgba(meta['color'], 0.75))

                brand_sub = sub[sub['Brand'] == brand]
                for grade, g_cnt in brand_sub['Sell Grade'].astype(str).str.strip().str.rstrip('.').value_counts().items():
                    g_id = f"lv_{code}_{brand}_{grade}"
                    lv_ids.append(g_id); lv_labels.append(f"{grade}<br>{int(g_cnt)}")
                    lv_parents.append(b_id); lv_values.append(int(g_cnt))
                    lv_clrs.append(grade_colors.get(grade, '#888888'))

        fig_lv = go.Figure(go.Treemap(
            ids=lv_ids, labels=lv_labels, parents=lv_parents,
            values=lv_values, marker=dict(colors=lv_clrs),
            branchvalues='total',
            hovertemplate='<b>%{label}</b><br>Devices: %{value}<br>Share of parent: %{percentParent:.1%}<extra></extra>',
            textfont=dict(size=12, color='white'),
            pathbar=dict(visible=True, thickness=20),
        ))
        fig_lv.update_layout(
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(t=30, b=10, l=10, r=10), height=460,
        )
        st.plotly_chart(fig_lv, use_container_width=True)
        st.caption("Click any block to drill in. Use the breadcrumb bar at top to navigate back.")

    with lv_tab2:
        st.markdown("**Download full data for each low value group**")
        lv_dl_cols = st.columns(3)
        for i, (code, meta) in enumerate(LOW_VALUE.items()):
            sub = get_subset(code).drop(columns=['_qr', '_cat'], errors='ignore')
            cache_key = f'_si_lv_{code}'
            if cache_key not in st.session_state:
                buf = BytesIO()
                import xlsxwriter as xw
                with xw.Workbook(buf, {'in_memory': True}) as wb:
                    ws = wb.add_worksheet(code)
                    hdr_fmt = wb.add_format({'bold': True, 'bg_color': '#4a235a', 'font_color': 'white', 'border': 1})
                    txt_fmt = wb.add_format({'num_format': '@'})
                    for ci, col_name in enumerate(sub.columns):
                        ws.write(0, ci, col_name, hdr_fmt)
                        ws.set_column(ci, ci, 20)
                    for ri, row_data in enumerate(sub.itertuples(index=False), start=1):
                        for ci, val in enumerate(row_data):
                            if sub.columns[ci] == 'IMEI Number':
                                ws.write_string(ri, ci, str(val).strip().split('.')[0] if pd.notna(val) else '', txt_fmt)
                            else:
                                ws.write(ri, ci, val if pd.notna(val) else '')
                buf.seek(0)
                st.session_state[cache_key] = buf.read()
            with lv_dl_cols[i]:
                st.download_button(
                    f"⬇️ {code}",
                    data=st.session_state[cache_key],
                    file_name=f"LowValue_{code}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                    key=f'dl_lv_{code}'
                )
                st.caption(f"{len(sub):,} rows")


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
# ── Session state init ────────────────────────────────────────────────────────
if 'show_analytics' not in st.session_state:
    st.session_state.show_analytics = False
if 'show_diagnostics' not in st.session_state:
    st.session_state.show_diagnostics = False
if 'show_delta' not in st.session_state:
    st.session_state.show_delta = False
if 'show_stock_intel' not in st.session_state:
    st.session_state.show_stock_intel = False
if 'master_df_result' not in st.session_state:
    st.session_state.master_df_result = None
if 'lv_master_df' not in st.session_state:
    st.session_state.lv_master_df = None

st.markdown(
    '<div class="main-header">📊 Master Template Generator</div>'
    '<div class="main-subtitle">Hanger + Totes + Stack Bulk → unified StockTake template</div>',
    unsafe_allow_html=True,
)
st.markdown("<p style='text-align:center; color:#666;'>Upload Hanger and/or Totes + Stack Bulk → Get Final Master Template</p>", unsafe_allow_html=True)
st.markdown("---")

# ── 3 upload buttons: Hanger | Totes (+ Low Value) | Stack Bulk ──────────────
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown(
        '<div class="upload-card">'
        '<div class="upload-title">📁 Hanger Stocktake File</div>'
        '<div class="upload-sub">Upload one or more hanger .xlsm files — all will be merged automatically</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    hanger_files = st.file_uploader(
        "Hanger file",
        type=['xlsx', 'xlsm', 'xls'],
        key="hanger",
        label_visibility="collapsed",
        accept_multiple_files=True,
    )
    hanger_file = hanger_files[0] if len(hanger_files) == 1 else (hanger_files if hanger_files else None)
    if hanger_files:
        for f in hanger_files:
            st.success(f"✅ {f.name}")

with col2:
    st.markdown(
        '<div class="upload-card">'
        '<div class="upload-title">🗂️ Totes / Low Value Stocktake File</div>'
        '<div class="upload-sub">Upload totes and/or low value .xlsm files — Low Value files are auto-detected and included in the master template</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    totes_files = st.file_uploader(
        "Totes file",
        type=['xlsx', 'xlsm', 'xls'],
        key="totes",
        label_visibility="collapsed",
        accept_multiple_files=True,
    )
    totes_file = totes_files[0] if len(totes_files) == 1 else (totes_files if totes_files else None)
    if totes_files:
        for f in totes_files:
            _label = "🏷️ Low Value" if is_low_value_file(f) else "🗂️ Totes"
            st.success(f"✅ {f.name} — {_label}")

with col3:
    st.markdown(
        '<div class="upload-card">'
        '<div class="upload-title">📋 Stack Bulk Upload File</div>'
        '<div class="upload-sub">Upload the Stack Bulk .xlsx file from the third party software</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    stack_file = st.file_uploader(
        "Stack Bulk file",
        type=['xlsx', 'xlsm', 'xls'],
        key="stack",
        label_visibility="collapsed"
    )
    if stack_file:
        st.success(f"✅ {stack_file.name}")

st.markdown("---")

# ── Validation messages ───────────────────────────────────────────────────────
if not stack_file:
    st.warning("⚠️ Stack Bulk Upload file is mandatory — please upload it to proceed.")

if stack_file and not hanger_files and not totes_files:
    st.warning("⚠️ Please upload at least one stocktake file — Hanger or Totes.")

# ── Generate button ───────────────────────────────────────────────────────────
has_stocktake = bool(hanger_files or totes_files)

if has_stocktake and stack_file:
    _run_main = st.button("🚀 Generate Master Template", type="primary", use_container_width=True)
    if _run_main:

        frames = []
        h_stats = None
        t_stats = None
        lv_frames = []   # Low Value rows — processed separately, merged at end

        if hanger_files:
            h_frames = []
            for _hf in hanger_files:
                with st.spinner(f"Processing hanger file: {_hf.name}..."):
                    try:
                        _hdf, _hs = process_hanger(_hf)
                        h_frames.append(_hdf)
                        if h_stats is None:
                            h_stats = dict(_hs)
                        else:
                            for k in h_stats:
                                h_stats[k] += _hs[k]
                    except Exception as e:
                        st.error(f"❌ Error reading {_hf.name}: {e}")
                        st.stop()
            if h_frames:
                frames.append(pd.concat(h_frames, ignore_index=True))

        if totes_files:
            t_frames = []
            for _tf in totes_files:
                if is_low_value_file(_tf):
                    with st.spinner(f"Processing Low Value file: {_tf.name}..."):
                        try:
                            _lv_df = build_low_value_template(_tf)
                            lv_frames.append(_lv_df)
                        except Exception as e:
                            st.error(f"❌ Error reading Low Value file {_tf.name}: {e}")
                            st.stop()
                else:
                    with st.spinner(f"Processing totes file: {_tf.name}..."):
                        try:
                            _tdf, _ts = process_totes(_tf)
                            t_frames.append(_tdf)
                            if t_stats is None:
                                t_stats = dict(_ts)
                            else:
                                for k in t_stats:
                                    t_stats[k] += _ts[k]
                        except Exception as e:
                            st.error(f"❌ Error reading {_tf.name}: {e}")
                            st.stop()
            if t_frames:
                frames.append(pd.concat(t_frames, ignore_index=True))

        with st.spinner("Reading Stack Bulk Upload..."):
            stack_df, stack_warn = read_stack_bulk(stack_file)
            if stack_df is None:
                st.error("❌ Could not read the Stack Bulk Upload file. Try opening it in Excel and saving as a new .xlsx file, then re-upload.")
                st.stop()
            if stack_warn and stack_warn != "ERROR":
                st.warning(f"⚠️ {stack_warn}")

        with st.spinner("Building lookup and generating master template..."):
            lookup = build_lookup(stack_df)
            if frames:
                combined_df = pd.concat(frames, ignore_index=True)
                master_df = build_master_template(combined_df, lookup)
            else:
                master_df = pd.DataFrame()

            # Merge Low Value rows into master template
            if lv_frames:
                lv_combined = pd.concat(lv_frames, ignore_index=True)

                # Remove completely blank rows (no IMEI and no Deal Id)
                lv_combined = lv_combined[
                    ~((lv_combined['IMEI'].astype(str).str.strip() == '') &
                      (lv_combined['Deal Id'].astype(str).str.strip().isin(['', 'nan', 'No deal ID'])))
                ].reset_index(drop=True)

                # Supplement with Stack Bulk lookup for LV rows missing Category
                empty_cat = lv_combined['Category'].astype(str).str.strip().isin(['', 'nan'])
                if empty_cat.any():
                    for idx in lv_combined[empty_cat].index:
                        deal_id = str(lv_combined.at[idx, 'Deal Id']).strip()
                        s = lookup.get(deal_id, {})
                        if s:
                            lv_combined.at[idx, 'Category'] = s.get('Category', '')
                            lv_combined.at[idx, 'Brand']    = s.get('Brand', '')
                            lv_combined.at[idx, 'Model']    = s.get('Model', '')
                            lv_combined.at[idx, 'Grade']    = s.get('Grade', '')
                            lv_combined.at[idx, 'VAT Type'] = s.get('VAT Type', '')
                            lv_combined.at[idx, 'Status']   = s.get('Status', '')
                            lv_combined.at[idx, 'Stack']    = s.get('Stack', '')

                master_df = pd.concat([master_df, lv_combined], ignore_index=True) if not master_df.empty else lv_combined

            st.session_state.master_df_result = master_df

        # ── Stats ─────────────────────────────────────────────────────────────
        st.markdown("### 📊 Processing Summary")

        lv_row_count = sum(len(f) for f in lv_frames) if lv_frames else 0

        if h_stats and t_stats:
            # Both hanger and totes uploaded — show side by side
            hc1, hc2, hc3, hc4 = st.columns(4)
            st.markdown("**Hanger**")
            with hc1:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#4472C4;">{h_stats["original_rows"]:,}</div><div class="metric-lbl">Hanger Rows</div></div>', unsafe_allow_html=True)
            with hc2:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#E74C3C;">{h_stats["ae_deal_id"]:,}</div><div class="metric-lbl">AE from Deal ID</div></div>', unsafe_allow_html=True)
            with hc3:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#70AD47;">{h_stats["ae_backend"]:,}</div><div class="metric-lbl">AE from BACKEND</div></div>', unsafe_allow_html=True)
            with hc4:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#E74C3C;">{h_stats["no_deal_id"]:,}</div><div class="metric-lbl">No Deal ID</div></div>', unsafe_allow_html=True)

            st.markdown("**Totes**")
            tc1, tc2, tc3, tc4 = st.columns(4)
            with tc1:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#4472C4;">{t_stats["original_rows"]:,}</div><div class="metric-lbl">Totes Rows</div></div>', unsafe_allow_html=True)
            with tc2:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#E74C3C;">{t_stats["ae_deal_id"]:,}</div><div class="metric-lbl">AE from Deal ID</div></div>', unsafe_allow_html=True)
            with tc3:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#70AD47;">{t_stats["ae_backend"]:,}</div><div class="metric-lbl">AE from BACKEND</div></div>', unsafe_allow_html=True)
            with tc4:
                st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#E74C3C;">{t_stats["no_deal_id"]:,}</div><div class="metric-lbl">No Deal ID</div></div>', unsafe_allow_html=True)

            if lv_row_count:
                st.markdown("**Low Value**")
                lvc1, _, _, _ = st.columns(4)
                with lvc1:
                    st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#9B59B6;">{lv_row_count:,}</div><div class="metric-lbl">Low Value Rows</div></div>', unsafe_allow_html=True)

        else:
            stats = h_stats or t_stats
            source = "Hanger" if h_stats else "Totes"
            if stats:
                c1, c2, c3, c4, c5 = st.columns(5)
                metrics = [
                    (c1, stats['original_rows'], f"Total {source} Rows", "#4472C4"),
                    (c2, stats['ae_deal_id'],    "AE from Deal ID",      "#E74C3C"),
                    (c3, stats['ae_backend'],    "AE from BACKEND",      "#70AD47"),
                    (c4, stats['no_deal_id'],    "No Deal ID",           "#E74C3C"),
                    (c5, len(stack_df),          "Stack Records",        "#8E44AD"),
                ]
                for col, val, lbl, color in metrics:
                    with col:
                        st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:{color};">{val:,}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

            elif lv_row_count:
                lvc1, _, _ = st.columns(3)
                with lvc1:
                    st.markdown(f'<div class="metric-box"><div class="metric-num" style="color:#9B59B6;">{lv_row_count:,}</div><div class="metric-lbl">Low Value Rows</div></div>', unsafe_allow_html=True)

        # ── Match stats ───────────────────────────────────────────────────────
        st.markdown("---")
        matched = master_df[
            (master_df['Deal Id'] != 'No deal ID') & (master_df['Category'] != '')
        ].shape[0]
        unmatched = master_df[
            (master_df['Deal Id'] != 'No deal ID') & (master_df['Category'] == '')
        ].shape[0]

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            st.markdown(f"""
            <div class="success-card">
                <strong>✅ Matched from Stack Bulk</strong><br>
                <span style="font-size:1.8rem; font-weight:700;">{matched:,}</span> rows
            </div>
            """, unsafe_allow_html=True)
        with mc2:
            st.markdown(f"""
            <div class="warn-card">
                <strong>⚠️ AE Deal ID — Not in Stack Bulk</strong><br>
                <span style="font-size:1.8rem; font-weight:700;">{unmatched:,}</span> rows
            </div>
            """, unsafe_allow_html=True)
        with mc3:
            st.markdown(f"""
            <div class="info-card">
                <strong>📋 Total Output Rows</strong><br>
                <span style="font-size:1.8rem; font-weight:700;">{len(master_df):,}</span> rows
            </div>
            """, unsafe_allow_html=True)

        # ── Preview ───────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 👀 Preview")
        tab1, tab2, tab3 = st.tabs(["✅ Matched Rows", "⚠️ Unmatched AE Rows", "📋 All Rows"])

        with tab1:
            matched_df = master_df[
                (master_df['Deal Id'] != 'No deal ID') & (master_df['Category'] != '')
            ]
            st.dataframe(matched_df.head(50), use_container_width=True)
            st.caption(f"Showing first 50 of {len(matched_df):,} matched rows")

        with tab2:
            unmatched_df = master_df[
                (master_df['Deal Id'] != 'No deal ID') & (master_df['Category'] == '')
            ]
            st.dataframe(unmatched_df.head(50), use_container_width=True)
            st.caption(f"Showing first 50 of {len(unmatched_df):,} rows — AE deal ID found but not in Stack Bulk")

        with tab3:
            st.dataframe(master_df.head(50), use_container_width=True)
            st.caption(f"Showing first 50 of {len(master_df):,} total rows")

        # ── Download ──────────────────────────────────────────────────────────
        st.markdown("---")
        with st.spinner("Preparing Excel file..."):
            excel_bytes = export_excel(master_df)

        filename = f"Master_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label=f"⬇️ Download Master Template — {len(master_df):,} rows",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

elif not has_stocktake or not stack_file:
    pass  # warnings already shown above

# ── Dashboard buttons — shown whenever hanger/totes is uploaded ──────────────
if has_stocktake:
    st.markdown("---")
    btn_row1_c1, btn_row1_c2 = st.columns(2)
    btn_row2_c1, btn_row2_c2 = st.columns(2)

    with btn_row1_c1:
        if st.button("📊 View Analytics Dashboard", use_container_width=True):
            st.session_state.show_analytics   = not st.session_state.show_analytics
            st.session_state.show_diagnostics = False
            st.session_state.show_delta       = False
            st.session_state.show_stock_intel = False

    with btn_row1_c2:
        if st.button("🔬 Smart Diagnostics", use_container_width=True):
            st.session_state.show_diagnostics = not st.session_state.show_diagnostics
            st.session_state.show_analytics   = False
            st.session_state.show_delta       = False
            st.session_state.show_stock_intel = False

    with btn_row2_c1:
        if st.button("📅 Daily Comparison", use_container_width=True):
            st.session_state.show_delta       = not st.session_state.show_delta
            st.session_state.show_analytics   = False
            st.session_state.show_diagnostics = False
            st.session_state.show_stock_intel = False

    with btn_row2_c2:
        if st.button("📦 Stock Intelligence", use_container_width=True):
            st.session_state.show_stock_intel = not st.session_state.show_stock_intel
            st.session_state.show_analytics   = False
            st.session_state.show_diagnostics = False
            st.session_state.show_delta       = False

    master_for_dash = st.session_state.get('master_df_result', None)

    _dash_hanger = hanger_files[0] if hanger_files else None
    # Only pass a regular totes file (not LV) to dashboards — LV files lack 'Room' column
    _regular_totes = [f for f in totes_files if not is_low_value_file(f)]
    _dash_totes = _regular_totes[0] if _regular_totes else None

    if st.session_state.show_analytics:
        show_analytics(_dash_hanger, _dash_totes, master_for_dash)

    if st.session_state.show_diagnostics:
        show_diagnostics(_dash_hanger, _dash_totes, master_for_dash)

    if st.session_state.show_delta:
        show_daily_delta(_dash_hanger, _dash_totes)

    if st.session_state.show_stock_intel:
        if not stack_file:
            st.warning("⚠️ Please upload a Stack Bulk file to view Stock Intelligence.")
        else:
            show_stock_intelligence(stack_file)

# Footer
st.markdown("---")
st.markdown(
    "<p style='text-align:center; color:#aaa; font-size:0.85rem;'>Operations Automation — Master Template Generator</p>",
    unsafe_allow_html=True
)
